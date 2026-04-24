"""
Flask web application for event revenue prediction using ensemble models.
Provides an interactive form for inputting event details and returns predictions with confidence intervals.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, make_response
from functools import wraps
from urllib.parse import urlparse, unquote
import joblib
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import json
import tempfile
import requests
import uuid
import hashlib
import re
from catboost import Pool
from weather_data import WeatherDataEnricher
from joblib import load as joblib_load

# Import EnsembleRegressor so joblib can unpickle it
try:
    from Kona_AI_ML.Kona_AI_ML import EnsembleRegressor
except ImportError:
    try:
        # Try direct import if running as script
        import sys
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        from Kona_AI_ML import EnsembleRegressor
    except ImportError:
        print("Warning: EnsembleRegressor not available. Ensemble models won't load.")
        EnsembleRegressor = None

# Import franchise database and auth
try:
    from Kona_AI_ML.franchise_db import FranchiseDatabase
    from Kona_AI_ML.auth import AuthManager, SessionManager
    from Kona_AI_ML.franchise_model_training import FranchiseModelTrainer
except ImportError:
    from franchise_db import FranchiseDatabase
    from auth import AuthManager, SessionManager
    from franchise_model_training import FranchiseModelTrainer

# Use a single source of truth for ZIP cluster prediction (DRY)
try:
    # Script-style first (when running from repo root)
    from zip_cluster_predict import predict_zip_cluster  # type: ignore
except Exception:
    # Package-style fallback
    from Kona_AI_ML.zip_cluster_predict import predict_zip_cluster  # type: ignore

# Import advanced prediction explainer
try:
    from Kona_AI_ML.prediction_explainer import PredictionExplainer
except ImportError:
    try:
        from prediction_explainer import PredictionExplainer
    except ImportError:
        print("Warning: PredictionExplainer not available. Advanced explanations will be disabled.")
        PredictionExplainer = None

try:
    from uszipcode import SearchEngine
    USZIPCODE_AVAILABLE = True
except ImportError:
    USZIPCODE_AVAILABLE = False
    SearchEngine = None
    print("Warning: uszipcode package not available. Install with: pip install uszipcode")

ENABLE_USZIPCODE = os.environ.get("ENABLE_USZIPCODE", "true").lower() == "true"
_zip_search_engine = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
TEMPLATES_DIR = os.path.join(PARENT_DIR, 'templates')

app = Flask(__name__, template_folder=TEMPLATES_DIR)

DATE_DISPLAY_FORMAT = "%m/%d/%Y"
REALIZED_EVENT_STATUSES = {'completed'}
PIPELINE_EVENT_STATUSES = {'booked_confirmed', 'needs_outcome'}
FORECAST_EVENT_STATUSES = {'predicted_only'}
DASHBOARD_PAGE_SIZE = 20
DASHBOARD_ALLOWED_STATUSES = {
    'predicted_only',
    'booked_confirmed',
    'needs_outcome',
    'completed',
    'cancelled',
    'rescheduled',
}
DASHBOARD_SORT_FIELDS = {
    'created_timestamp',
    'event_name',
    'event_status',
    'event_date',
    'predicted_revenue_per_hour',
    'predicted_total_revenue',
    'actual_total_net_sales',
    'predicted_vs_actual_diff',
    'actual_revenue_per_hour',
    'confidence_lower',
    'confidence_upper',
    'duration_hours',
}


def _normalize_event_status(value, default='predicted_only'):
    allowed_statuses = {
        'predicted_only',
        'booked_confirmed',
        'needs_outcome',
        'completed',
        'cancelled',
        'rescheduled',
    }
    status = str(value or '').strip().lower()
    return status if status in allowed_statuses else default


def _parse_positive_int(value, default=1):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _get_dashboard_state(values):
    page = _parse_positive_int(values.get('page'), default=1)
    month_key = _normalize_dashboard_month(values.get('month'))

    status_filter = str(values.get('status') or 'all').strip().lower()
    if status_filter != 'all' and status_filter not in DASHBOARD_ALLOWED_STATUSES:
        status_filter = 'all'

    sort_by = str(values.get('sort_by') or 'created_timestamp').strip().lower()
    if sort_by not in DASHBOARD_SORT_FIELDS:
        sort_by = 'created_timestamp'

    sort_dir = str(values.get('sort_dir') or 'desc').strip().lower()
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    return {
        'page': page,
        'month': month_key,
        'status': status_filter,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
    }


def _normalize_dashboard_month(value):
    default_month = datetime.now().strftime("%Y-%m")
    month_key = str(value or default_month).strip()
    try:
        datetime.strptime(month_key, "%Y-%m")
        return month_key
    except ValueError:
        return default_month


def _dashboard_month_bounds(month_key):
    first_of_month = datetime.strptime(month_key, "%Y-%m")
    year = first_of_month.year
    month = first_of_month.month

    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)

    if month == 1:
        prev_month = datetime(year - 1, 12, 1)
    else:
        prev_month = datetime(year, month - 1, 1)

    return {
        'start': first_of_month.date(),
        'next_start': next_month.date(),
        'label': first_of_month.strftime("%B %Y"),
        'prev_key': prev_month.strftime("%Y-%m"),
        'next_key': next_month.strftime("%Y-%m"),
    }


def _try_parse_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        normalized = value.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            pass

        formats = [
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%m/%d/%Y",
            "%m/%d/%Y %H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


@app.template_filter("mdy_date")
def mdy_date(value):
    parsed = _try_parse_datetime(value)
    if parsed is None:
        return value or ""
    return parsed.strftime(DATE_DISPLAY_FORMAT)


@app.template_filter("compact_currency")
def compact_currency(value):
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return "$0"

    sign = "-" if amount < 0 else ""
    absolute = abs(amount)

    if absolute >= 1_000_000_000:
        return f"{sign}${absolute / 1_000_000_000:.2f}B"
    if absolute >= 1_000_000:
        return f"{sign}${absolute / 1_000_000:.2f}M"
    if absolute >= 1_000:
        return f"{sign}${absolute / 1_000:.1f}K"
    return f"{sign}${absolute:,.2f}"

# Security configuration for sessions
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', 'false').lower() == 'true'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Get the script directory to make paths relative to it

# Initialize database and auth
DB_PATH = os.path.join(SCRIPT_DIR, 'franchise_data.db')

franchise_db = FranchiseDatabase(DB_PATH)
auth_manager = AuthManager(franchise_db)

print(f"Database initialized at: {DB_PATH}")

# ===== AUTHENTICATION HELPERS =====

def require_auth(f):
    """Decorator to require authentication for protected routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check for session cookie
        session_token = request.cookies.get(SessionManager.get_session_cookie_name())
        if not session_token:
            return redirect(url_for('login_page', next=request.url))
        
        # Validate session
        franchise_id, error = auth_manager.validate_session(session_token)
        if not franchise_id:
            # Session expired or invalid, redirect to login
            response = make_response(redirect(url_for('login_page', next=request.url)))
            response.delete_cookie(SessionManager.get_session_cookie_name())
            return response
        
        # Store franchise info in request context for use in route handler
        request.franchise_id = franchise_id
        request.franchise = franchise_db.get_franchise(franchise_id)
        
        return f(*args, **kwargs)
    
    return decorated_function

def get_current_franchise():
    """Get current franchise from request context."""
    return request.franchise

def get_current_franchise_id():
    """Get current franchise ID from request context."""
    return request.franchise_id

def is_safe_redirect_target(target: str) -> bool:
    """
    Return True only if `target` is a safe redirect URL:
      - A relative path starting with '/' and NOT starting with '//' (protocol-relative)
      - OR an absolute URL whose scheme is http/https and whose host matches the current request host
    Normalizes backslashes and URL-encoded characters to catch common bypass attempts.
    """
    if not target:
        return False
    # Decode URL-encoded characters (e.g. %5C -> \) then normalize backslashes
    target = unquote(target).replace('\\', '/')
    parsed = urlparse(target)
    # Relative path: no scheme, no netloc, must start with /
    if not parsed.scheme and not parsed.netloc:
        return target.startswith('/') and not target.startswith('//')
    # Absolute URL: scheme must be http/https and host must match current host
    if parsed.scheme in ('http', 'https'):
        return parsed.netloc == request.host
    return False

# ===== MODEL LOADING =====

prod_models_dir = os.path.join(PARENT_DIR, 'production', 'models')

# Load clean_layers production ensemble via ProductionModelManager
try:
    from production_model_manager import ProductionModelManager
    
    candidate_model_dirs = [
        os.path.join(PARENT_DIR, 'production', 'models'),
        os.path.join(PARENT_DIR, 'Kona_AI_ML', 'production', 'models'),
    ]
    prod_models_dir = next((path for path in candidate_model_dirs if os.path.isdir(path)), candidate_model_dirs[0])
    PROD_MANAGER = ProductionModelManager(models_dir=prod_models_dir)
    
    # Load metrics from production deployment
    metrics_path = os.path.join(prod_models_dir, 'clean_layers_training_metrics.json')
    with open(metrics_path, 'r') as f:
        metrics_data = json.load(f)
        cv_mae = metrics_data.get('model_scores', {}).get('xgboost', {}).get('MAE', 312.66)
        cv_std = 50  # Approximate from training
    
    model = PROD_MANAGER
    MODEL_PATH = prod_models_dir
    print(f"✓ Loaded clean_layers production ensemble from {prod_models_dir}")
    print(f"  Models: XGBoost, CatBoost, LightGBM (ensemble)")
    
except Exception as e:
    print(f"ERROR: Could not load production models: {e}")
    import traceback
    traceback.print_exc()
    PROD_MANAGER = None
    model = None
    MODEL_PATH = None
    cv_mae = 312.66
    cv_std = 50

FRANCHISE_MODEL_TRAINER = None
if PROD_MANAGER is not None:
    try:
        FRANCHISE_MODEL_TRAINER = FranchiseModelTrainer(
            franchise_db,
            prod_models_dir,
            threshold=PROD_MANAGER.FRANCHISE_DATA_THRESHOLD,
            min_r2=PROD_MANAGER.FRANCHISE_MODEL_CONFIDENCE_R2,
        )
    except Exception as trainer_error:
        print(f"[WARN] Franchise model trainer unavailable: {trainer_error}")
if cv_mae is not None and cv_std is not None:
    try:
        print(f"CV MAE: {float(cv_mae):.2f} ± {float(cv_std):.2f}")
    except Exception:
        print(f"CV metrics found but not printable: mae={cv_mae}, std={cv_std}")
else:
    print("CV metrics not available; using conformal intervals if supported.")

# Initialize advanced prediction explainer lazily to reduce baseline RAM usage
ENABLE_PREDICTION_EXPLAINER = os.environ.get("ENABLE_PREDICTION_EXPLAINER", "true").lower() == "true"
PREDICTION_EXPLAINER = None


def _get_zip_search_engine():
    global _zip_search_engine
    if not ENABLE_USZIPCODE or not USZIPCODE_AVAILABLE:
        return None
    if _zip_search_engine is None:
        try:
            _zip_search_engine = SearchEngine()
        except Exception as e:
            print(f"Warning: uszipcode SearchEngine init failed: {e}")
            _zip_search_engine = None
    return _zip_search_engine


def get_prediction_explainer():
    global PREDICTION_EXPLAINER
    if not ENABLE_PREDICTION_EXPLAINER or PredictionExplainer is None:
        return None
    if PREDICTION_EXPLAINER is None:
        try:
            print("Initializing advanced prediction explainer...")
            # Note: clean_layers model generates features at runtime, no static background data needed
            PREDICTION_EXPLAINER = PredictionExplainer(
                model,
                X_background=None,
                use_shap=True
            )
            print("[OK] Prediction explainer initialized successfully")
        except Exception as e:
            print(f"[WARN] Could not initialize prediction explainer: {e}")
            PREDICTION_EXPLAINER = None
    return PREDICTION_EXPLAINER

# With clean_layers architecture, features are generated at runtime
# No need to load template or historical data for prediction
print("Clean layers architecture: features generated at runtime")

# Load Events-in-County counts lookup if available
EVENTS_IN_COUNTY_LOOKUP = None
EVENTS_IN_COUNTY_PATH = os.path.join(SCRIPT_DIR, 'outputs', 'events_in_county_counts.csv')
if os.path.exists(EVENTS_IN_COUNTY_PATH):
    try:
        _eic_df = pd.read_csv(EVENTS_IN_COUNTY_PATH)
        # Normalize to dict with key (county, date)
        _eic_df['County'] = _eic_df['County'].astype(str).str.strip()
        _eic_df['Date'] = _eic_df['Date'].astype(str)
        EVENTS_IN_COUNTY_LOOKUP = {
            (row['County'], row['Date']): int(row['Count'])
            for _, row in _eic_df.iterrows()
        }
        print(f"Events-in-County lookup loaded: {len(EVENTS_IN_COUNTY_LOOKUP)} keys")
    except Exception as e:
        print(f"Warning: Failed to load events-in-county lookup: {e}")

# Clean layers model uses 46 features generated at runtime via clean_layers_feature_preparation
print(f"Clean layers model: 46 features generated at prediction time")

# Load feature importance data for explanations
FEATURE_IMPORTANCE_DF = None
FEATURE_IMPORTANCE_DICT = {}
# Try both locations: parent_dir/outputs and SCRIPT_DIR/outputs
possible_paths = [
    os.path.join(PARENT_DIR, 'outputs', 'feature_importances.csv'),  # Root level outputs/
    os.path.join(SCRIPT_DIR, 'outputs', 'feature_importances.csv'),   # Kona_AI_ML/outputs/
]
FEATURE_IMPORTANCE_PATH = None
for path in possible_paths:
    if os.path.exists(path):
        FEATURE_IMPORTANCE_PATH = path
        break

if FEATURE_IMPORTANCE_PATH:
    try:
        FEATURE_IMPORTANCE_DF = pd.read_csv(FEATURE_IMPORTANCE_PATH)
        FEATURE_IMPORTANCE_DICT = dict(zip(FEATURE_IMPORTANCE_DF['feature'], FEATURE_IMPORTANCE_DF['importance']))
        print(f"Feature importance data loaded: {len(FEATURE_IMPORTANCE_DICT)} features from {FEATURE_IMPORTANCE_PATH}")
    except Exception as e:
        print(f"[WARN] Could not load feature importances: {e}")
        FEATURE_IMPORTANCE_DF = pd.DataFrame()  # Empty dataframe as fallback
else:
    print(f"[WARN] Feature importance file not found in: {possible_paths[0]} or {possible_paths[1]}")
    FEATURE_IMPORTANCE_DF = pd.DataFrame()  # Empty dataframe as fallback

# Cache for zip code lookups to avoid repeated API calls
ZIP_CACHE = {}

# Initialize weather data enricher
print("Initializing weather data provider...")
WEATHER_ENRICHER = WeatherDataEnricher(provider='open-meteo')
print(f"Weather provider initialized: {WEATHER_ENRICHER.weather_provider is not None}")


# ===== AVAILABLE MODELS DETECTION =====

def _get_prod_models_dir():
    return MODEL_PATH or os.path.join(PARENT_DIR, 'production', 'models')


def _serialize_feature_snapshot(feature_df):
    """Serialize a single-row model feature frame for future franchise retraining."""
    if feature_df is None or getattr(feature_df, 'empty', True):
        return None

    snapshot = {}
    for key, value in feature_df.iloc[0].to_dict().items():
        if pd.isna(value):
            snapshot[key] = 0.0
        elif isinstance(value, (np.integer, int)):
            snapshot[key] = int(value)
        elif isinstance(value, (np.floating, float)):
            snapshot[key] = float(value)
        elif isinstance(value, (np.bool_, bool)):
            snapshot[key] = int(value)
        else:
            snapshot[key] = value
    return json.dumps(snapshot)


def _parse_training_metadata(raw_metadata):
    if not raw_metadata:
        return {}
    try:
        return json.loads(raw_metadata)
    except (TypeError, ValueError):
        return {}

def get_available_models():
    """
    Discover available models from the production deployment.
    Returns models based on actual files in production/models/ and deployment metadata.
    """
    models = []
    
    # Try to load deployment metadata first
    metadata_path = os.path.join(_get_prod_models_dir(), 'deployment_metadata.json')
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, 'r') as f:
                metadata = json.load(f)
            
            # Add ensemble models from metadata
            deployment_type = metadata.get('deployment_type', 'clean_layers_ensemble')
            perf = metadata.get('performance_metrics', {})
            
            # Clean layers ensemble
            if deployment_type == 'clean_layers_ensemble':
                models.append({
                    'model_id': 'clean_layers_ensemble',
                    'model_name': 'Clean Layers Ensemble (XGBoost + CatBoost + LightGBM)',
                    'model_type': 'ensemble',
                    'description': f"Merged ensemble model (46 features, 3 sub-models)",
                    'mae': perf.get('ensemble_mae', 312.66),
                    'cv_mae': perf.get('ensemble_mae', 312.66),
                    'r2': perf.get('ensemble_r2', 0.305),
                    'training_date': metadata.get('deployment_timestamp', ''),
                    'feature_count': metadata.get('training_metadata', {}).get('total_features', 46),
                    'status': 'production',
                    'selectable': True,
                })
                
                # Add individual ensemble models
                for model_name in ['xgboost', 'catboost', 'lightgbm']:
                    model_perf = perf.get('individual_models', {}).get(model_name, {})
                    models.append({
                        'model_id': f'clean_layers_{model_name}',
                        'model_name': f'Clean Layers - {model_name.upper()}',
                        'model_type': 'individual',
                        'description': f"Individual {model_name.upper()} model (component of ensemble)",
                        'mae': model_perf.get('MAE', 0),
                        'cv_mae': model_perf.get('MAE', 0),
                        'r2': model_perf.get('R2', 0),
                        'training_date': metadata.get('deployment_timestamp', ''),
                        'feature_count': metadata.get('training_metadata', {}).get('total_features', 46),
                        'status': 'production',
                        'selectable': True,
                    })
        
        except Exception as e:
            print(f"[WARN] Could not load deployment metadata: {e}")
    
    # Check for standalone models in production directory
    prod_dir = _get_prod_models_dir()
    if os.path.isdir(prod_dir):
        # Look for standalone model files
        for filename in os.listdir(prod_dir):
            if filename.endswith('_model.joblib') and not filename.startswith('franchise_'):
                model_name = filename.replace('_model.joblib', '')
                
                # Skip if already added
                if any(m['model_id'] == model_name for m in models):
                    continue
                
                # Add standalone model
                models.append({
                    'model_id': model_name,
                    'model_name': f'{model_name.replace("_", " ").title()}',
                    'model_type': 'standalone',
                    'description': 'Standalone production model',
                    'mae': None,
                    'cv_mae': None,
                    'r2': None,
                    'training_date': '',
                    'feature_count': None,
                    'status': 'production',
                    'selectable': True,
                })
    
    return models


def get_franchise_available_models(franchise_id, franchise_progress=None):
    """
    Get available models for a franchise (merged + franchise-specific if exists).
    """
    available_models = get_available_models()
    progress = franchise_progress or franchise_db.get_franchise_model_progress(
        franchise_id,
        threshold=(PROD_MANAGER.FRANCHISE_DATA_THRESHOLD if PROD_MANAGER else 100),
    )
    latest_model = progress.get('latest_model')

    if latest_model:
        metadata = _parse_training_metadata(latest_model.get('training_metadata_json'))
        metrics = metadata.get('metrics', {})
        history = metadata.get('training_history') if isinstance(metadata, dict) else None
        if not isinstance(history, list):
            history = []
        if not history and metrics:
            history = [{
                'trained_at': metadata.get('trained_at'),
                'r2': metrics.get('r2'),
                'cv_mae_per_hour': metrics.get('cv_mae_per_hour') or latest_model.get('cv_mae'),
                'trainable_event_count': metadata.get('trainable_event_count') or latest_model.get('data_records_count'),
            }]

        next_retrain_event_count = progress.get('next_retrain_event_count')
        remaining_to_next_retrain = progress.get('remaining_to_next_retrain', 0)
        eligible_event_count = progress.get('eligible_event_count', 0)
        last_trained_event_count = progress.get('last_trained_event_count') or latest_model.get('data_records_count') or 0
        retrain_goal = max(1, (next_retrain_event_count or (last_trained_event_count + 1)) - last_trained_event_count)
        retrain_progress_count = max(0, eligible_event_count - last_trained_event_count)
        retrain_progress_pct = min(100, (retrain_progress_count / retrain_goal) * 100)
        available_models.append({
            'model_id': f'franchise_{franchise_id}',
            'model_name': 'Your Franchise Model',
            'model_type': 'franchise_specific',
            'description': (
                f"Auto-refreshes every milestone; next refresh at {next_retrain_event_count} completed events."
                if next_retrain_event_count else
                'Custom-trained model using your completed event history.'
            ),
            'mae': metrics.get('cv_mae_total'),
            'cv_mae': latest_model.get('cv_mae'),
            'cv_std': latest_model.get('cv_std'),
            'r2': metrics.get('r2', latest_model.get('r2')),
            'training_date': latest_model.get('training_date', ''),
            'feature_count': latest_model.get('feature_count', 46),
            'data_records_count': latest_model.get('data_records_count'),
            'status': 'production',
            'selectable': True,
            'progress_pct': retrain_progress_pct,
            'progress_event_count': retrain_progress_count,
            'progress_event_goal': retrain_goal,
            'progress_event_label': 'events toward next refresh',
            'eligible_event_count': eligible_event_count,
            'threshold_event_count': progress.get('threshold_event_count', 100),
            'remaining_events': 0,
            'next_retrain_event_count': next_retrain_event_count,
            'remaining_to_next_retrain': remaining_to_next_retrain,
            'last_trained_event_count': last_trained_event_count,
            'training_metrics': metrics,
            'last_training_message': progress.get('last_training_message'),
            'training_history': history,
        })
    else:
        ready_with_features = progress.get('ready_with_features', False)
        threshold_event_count = progress.get('threshold_event_count', 100)
        eligible_event_count = progress.get('eligible_event_count', 0)
        trainable_event_count = progress.get('trainable_event_count', 0)
        if progress.get('ready_for_training') and ready_with_features:
            status = 'training'
            description = 'Threshold reached. We are packaging your dedicated model now.'
            progress_pct = min(100, (trainable_event_count / max(1, threshold_event_count)) * 100)
            remaining_events = max(0, threshold_event_count - trainable_event_count)
            progress_event_count = trainable_event_count
            progress_event_label = 'feature-complete completed events'
        elif progress.get('ready_for_training'):
            status = 'calibrating'
            missing = max(0, threshold_event_count - trainable_event_count)
            description = (
                f'Threshold reached with {eligible_event_count} completed events, but only '
                f'{trainable_event_count} include feature snapshots required for training. '
                f'We need {missing} more feature-complete completed events.'
            )
            progress_pct = min(100, (trainable_event_count / max(1, threshold_event_count)) * 100)
            remaining_events = missing
            progress_event_count = trainable_event_count
            progress_event_label = 'feature-complete completed events'
        else:
            status = 'locked'
            description = 'Unlocks after enough completed events are recorded with actual outcomes.'
            progress_pct = progress.get('progress_pct', 0)
            remaining_events = progress.get('remaining_events', 0)
            progress_event_count = eligible_event_count
            progress_event_label = 'completed events'

        available_models.append({
            'model_id': f'franchise_{franchise_id}',
            'model_name': 'Your Franchise Model',
            'model_type': 'franchise_specific',
            'description': description,
            'mae': None,
            'cv_mae': None,
            'cv_std': None,
            'r2': None,
            'training_date': '',
            'feature_count': 46,
            'data_records_count': progress.get('trainable_event_count'),
            'status': status,
            'selectable': False,
            'progress_pct': progress_pct,
            'progress_event_count': progress_event_count,
            'progress_event_goal': threshold_event_count,
            'progress_event_label': progress_event_label,
            'eligible_event_count': eligible_event_count,
            'threshold_event_count': threshold_event_count,
            'remaining_events': remaining_events,
            'ready_for_training': progress.get('ready_for_training', False),
            'ready_with_features': ready_with_features,
            'trainable_event_count': trainable_event_count,
            'expected_ready_date': progress.get('expected_ready_date'),
            'last_training_status': progress.get('last_training_status'),
            'last_training_message': progress.get('last_training_message'),
        })

    # Sort selectable models by R² descending, cv_mae ascending; non-selectable models go to the bottom
    selectable = [m for m in available_models if m.get('selectable', True)]
    non_selectable = [m for m in available_models if not m.get('selectable', True)]
    selectable.sort(key=lambda m: (-(m.get('r2') or 0), (m.get('cv_mae') or float('inf'))))
    if selectable:
        selectable[0]['is_recommended'] = True
    available_models = selectable + non_selectable

    return available_models


def get_demographics_from_zip(zip_code):
    """
    Fetch demographic data from Census API for a given ZIP code.
    Falls back to Illinois averages if API fails.
    """
    zip_str = str(zip_code).zfill(5)
    
    # Check cache first
    if zip_str in ZIP_CACHE:
        return ZIP_CACHE[zip_str]
    
    try:
        # Census API endpoint for ZIP Code Tabulation Areas (ZCTA)
        # Using ACS 5-Year estimates for better coverage
        base_url = "https://api.census.gov/data/2021/acs/acs5"
        
        # Variables we need:
        # B01003_001E: Total Population
        # B19013_001E: Median Household Income
        # B23025_005E: Unemployment (in labor force, unemployed)
        # B23025_003E: In labor force
        # B25064_001E: Median Gross Rent
        # B25077_001E: Median Home Value
        # B11001_001E: Total Households
        # B11003_003E: Households with own children
        # B25010_001E: Average Household Size
        
        params = {
            'get': 'B01003_001E,B19013_001E,B23025_005E,B23025_003E,B25064_001E,B25077_001E,B11001_001E,B11003_003E,B25010_001E',
            'for': f'zip code tabulation area:{zip_str}',
        }
        
        response = requests.get(base_url, params=params, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            if len(data) > 1:  # First row is headers
                values = data[1]
                
                population = float(values[0]) if values[0] not in [None, '-666666666'] else 50000
                median_income = float(values[1]) if values[1] not in [None, '-666666666'] else 75000
                unemployed = float(values[2]) if values[2] not in [None, '-666666666'] else 0
                labor_force = float(values[3]) if values[3] not in [None, '-666666666'] else 1
                median_rent = float(values[4]) if values[4] not in [None, '-666666666'] else 1200
                median_home_value = float(values[5]) if values[5] not in [None, '-666666666'] else 300000
                total_households = float(values[6]) if values[6] not in [None, '-666666666'] else 20000
                households_with_children = float(values[7]) if values[7] not in [None, '-666666666'] else 5000
                avg_household_size = float(values[8]) if values[8] not in [None, '-666666666'] else 2.5
                
                # Calculate unemployment rate
                unemployment_rate = (unemployed / labor_force * 100) if labor_force > 0 else 4.0
                
                # Calculate rent to income ratio
                rent_to_income = (median_rent * 12 / median_income) if median_income > 0 else 0.3
                
                demographics = {
                    'Total_Population': population,
                    'Median_Household_Income': median_income,
                    'Unemployment_Rate': unemployment_rate,
                    'Average_Household_Size': avg_household_size,
                    'Rent_To_Income_Ratio': rent_to_income,
                    'Median_Rent': median_rent,
                    'Median_Home_Value': median_home_value,
                    'Households_With_Children': households_with_children,
                    'source': 'census_api'
                }
                
                ZIP_CACHE[zip_str] = demographics
                return demographics
    
    except Exception as e:
        print(f"Census API error for {zip_str}: {str(e)}")
    
    # Fallback to Illinois/DuPage County averages
    demographics = {
        'Total_Population': 50000,
        'Median_Household_Income': 75000,
        'Unemployment_Rate': 4.5,
        'Average_Household_Size': 2.6,
        'Rent_To_Income_Ratio': 0.30,
        'Median_Rent': 1400,
        'Median_Home_Value': 350000,
        'Households_With_Children': 15000,
        'source': 'default'
    }
    
    ZIP_CACHE[zip_str] = demographics
    return demographics


def get_county_from_zip(zip_code):
    """
    Get county name from ZIP code using Census API.
    Falls back to DuPage County if lookup fails.
    """
    zip_str = str(zip_code).zfill(5)
    
    try:
        # Use geocoding API to get county
        url = f"https://geocoding.geo.census.gov/geocoder/geographies/address"
        params = {
            'zip': zip_str,
            'benchmark': 'Public_AR_Current',
            'vintage': 'Current_Current',
            'format': 'json'
        }
        
        response = requests.get(url, params=params, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if 'result' in data and 'geographies' in data['result']:
                counties = data['result']['geographies'].get('Counties', [])
                if counties:
                    return counties[0]['NAME']
    
    except Exception as e:
        print(f"County lookup error for {zip_str}: {str(e)}")
    
    # Fallback
    return "DuPage County"


def get_seasonal_weather_estimate(event_datetime):
    """
    Get estimated weather based on season and month for Chicago/Illinois area.
    Provides more realistic estimates than static 70F default.
    
    Args:
        event_datetime: Event datetime
        
    Returns:
        dict: Estimated weather with temperature and rain probability
    """
    month = event_datetime.month
    
    # Chicago area seasonal temperature averages (°F) and rain probabilities
    seasonal_data = {
        1: {'temp_range': (25, 35), 'rain_prob': 0.25, 'season': 'Winter'},   # January
        2: {'temp_range': (29, 40), 'rain_prob': 0.23, 'season': 'Winter'},   # February  
        3: {'temp_range': (39, 52), 'rain_prob': 0.30, 'season': 'Spring'},   # March
        4: {'temp_range': (50, 65), 'rain_prob': 0.35, 'season': 'Spring'},   # April
        5: {'temp_range': (60, 75), 'rain_prob': 0.35, 'season': 'Spring'},   # May
        6: {'temp_range': (70, 82), 'rain_prob': 0.30, 'season': 'Summer'},   # June
        7: {'temp_range': (75, 85), 'rain_prob': 0.28, 'season': 'Summer'},   # July
        8: {'temp_range': (73, 83), 'rain_prob': 0.30, 'season': 'Summer'},   # August
        9: {'temp_range': (65, 76), 'rain_prob': 0.28, 'season': 'Fall'},     # September
        10: {'temp_range': (53, 66), 'rain_prob': 0.25, 'season': 'Fall'},    # October
        11: {'temp_range': (40, 53), 'rain_prob': 0.30, 'season': 'Fall'},    # November
        12: {'temp_range': (29, 40), 'rain_prob': 0.27, 'season': 'Winter'}   # December
    }
    
    month_data = seasonal_data[month]
    
    # Calculate average temperature for the month
    avg_temp = sum(month_data['temp_range']) / 2
    
    # Determine if it's likely to be raining (simplified: rain if probability > 30%)
    is_likely_raining = month_data['rain_prob'] > 0.30
    
    return {
        'temperature_f': round(avg_temp, 1),
        'is_raining': is_likely_raining,
        'source': 'seasonal_estimate',
        'season': month_data['season'],
        'rain_probability': round(month_data['rain_prob'] * 100, 1)
    }


def get_weather_source_description(weather_info):
    """Get a human-readable description of the weather data source."""
    if not weather_info:
        return "No weather data available"
    
    source = weather_info.get('source', 'unknown')
    descriptions = {
        'user_input': 'User provided',
        'weather_api': 'Historical weather data',
        'seasonal_estimate': f"Seasonal estimate ({weather_info.get('season', 'Unknown')} average)",
        'default': 'Default fallback'
    }
    return descriptions.get(source, 'Unknown source')


def get_weather_for_event(zip_code, event_datetime, user_temperature=None):
    """
    Get weather data for an event, either from user input or weather API.
    
    Args:
        zip_code: Event ZIP code
        event_datetime: Event start datetime
        user_temperature: User-provided temperature (optional)
        
    Returns:
        dict: Weather information including temperature and rain status
    """
    weather_info = {
        'temperature_f': 70.0,  # Default fallback
        'is_raining': False,
        'source': 'default'
    }
    
    # If user provided temperature, use it
    if user_temperature is not None:
        try:
            temp_f = float(user_temperature)
            weather_info['temperature_f'] = temp_f
            weather_info['source'] = 'user_input'
            weather_info['is_raining'] = False  # Assume no rain if user didn't specify
            return weather_info
        except (ValueError, TypeError):
            print(f"Invalid user temperature: {user_temperature}, falling back to API")
    
    # Try to get weather from API for historical data only
    if WEATHER_ENRICHER and WEATHER_ENRICHER.weather_provider:
        try:
            # Only use weather API for dates in the past
            from datetime import datetime as dt_check
            if event_datetime.date() < dt_check.now().date():
                # Get approximate coordinates for ZIP code
                lat, lon = get_approximate_coordinates(zip_code)
                
                temp_f, is_raining = WEATHER_ENRICHER.weather_provider.get_historical_weather(
                    lat, lon, event_datetime
                )
                
                if temp_f is not None:
                    weather_info['temperature_f'] = temp_f
                    weather_info['source'] = 'weather_api'
                
                if is_raining is not None:
                    weather_info['is_raining'] = is_raining
            else:
                # Future date - use seasonal defaults
                month = event_datetime.month
                if month in [12, 1, 2]:  # Winter
                    weather_info['temperature_f'] = 35.0
                elif month in [3, 4, 5]:  # Spring
                    weather_info['temperature_f'] = 60.0
                elif month in [6, 7, 8]:  # Summer
                    weather_info['temperature_f'] = 80.0
                else:  # Fall
                    weather_info['temperature_f'] = 55.0
                weather_info['source'] = 'seasonal_default'
            
            print(f"Weather API result for {zip_code}: {weather_info['temperature_f']}°F, rain={weather_info['is_raining']}")
            
        except Exception as e:
            print(f"Weather API error: {e}")
            # Use seasonal estimate instead of static default
            seasonal_weather = get_seasonal_weather_estimate(event_datetime)
            weather_info.update(seasonal_weather)
            print(f"Using seasonal estimate: {seasonal_weather['temperature_f']}°F, {seasonal_weather['season']}")
    else:
        # No weather API available, use seasonal estimate
        seasonal_weather = get_seasonal_weather_estimate(event_datetime)
        weather_info.update(seasonal_weather)
        print(f"No weather API, using seasonal estimate: {seasonal_weather['temperature_f']}°F, {seasonal_weather['season']}")
    
    return weather_info


def get_approximate_coordinates(zip_code):
    """
    Get accurate coordinates for a ZIP code using uszipcode database.
    Falls back to Chicago area coordinates if lookup fails.
    
    Args:
        zip_code: US ZIP code (string or int)
    
    Returns:
        Tuple of (latitude, longitude)
    """
    # Try uszipcode first
    zip_engine = _get_zip_search_engine()
    if zip_engine:
        try:
            zipcode_obj = zip_engine.by_zipcode(str(zip_code))
            if zipcode_obj and zipcode_obj.lat and zipcode_obj.lng:
                print(f"Found coordinates for ZIP {zip_code}: {zipcode_obj.lat}, {zipcode_obj.lng}")
                return float(zipcode_obj.lat), float(zipcode_obj.lng)
        except Exception as e:
            print(f"Error looking up ZIP {zip_code}: {e}")
    
    # Fallback to Chicago area coordinates
    print(f"Using fallback coordinates for ZIP {zip_code}")
    return 41.8781, -87.6298


## predict_zip_cluster is imported from zip_cluster_predict (DRY)


def get_events_in_county_count(zip_code: str, event_datetime: datetime) -> int:
    """
    Lookup historical number of events in the same county on the same date.
    Falls back to 1 if the lookup data or county is unknown.
    """
    # Default fallback
    default_val = 1
    try:
        date_key = event_datetime.date().isoformat()
        # Get county via demographics cache
        try:
            from demographics import get_zip_demographics
        except Exception:
            from Kona_AI_ML.demographics import get_zip_demographics  # type: ignore
        demo = get_zip_demographics(str(zip_code))
        county = demo.get('county') or demo.get('County')
        if not isinstance(county, str) or not county.strip():
            return default_val
        county = county.strip()
        if EVENTS_IN_COUNTY_LOOKUP is None:
            return default_val
        return int(EVENTS_IN_COUNTY_LOOKUP.get((county, date_key), default_val))
    except Exception as e:
        print(f"Events-in-County lookup error: {e}")
        return default_val

# Standard event industries for prediction form
INDUSTRIES = [
    'Community',
    'Athletics',
    'Corporate',
    'Education',
    'Festival',
    'Church',
    'Private',
    'Daycare'
]

# Feature list based on Phase 1 + Phase 2 features
# These will be populated from the form
CATEGORICAL_FEATURES = [
    'Industry', 'Season', 'Time of Day', 'Industry Category', 'Income_Tier', 'Population_Tier',
    'County_Industry', 'City_Industry', 'ZIP_Cluster_Industry_Category',
    'Season_Industry', 'TimeOfDay_Industry', 'Weekend_Industry', 'Month_Income_Tier',
    'Duration_Bracket', 'Duration_Bracket_Industry', 'Duration_TimeOfDay',
    'Income_Tier_Industry', 'Household_Size_Category', 'Household_Size_Industry',
    'School_Season', 'Hour_Bucket', 'Price_Tier_By_Industry',
    'ZIP Cluster'  # This was treated as categorical during training
]


def calculate_confidence_interval(prediction, mae, std):
    """
    Calculate confidence interval based on CV MAE and standard deviation.
    Returns lower and upper bounds, and a confidence percentage.
    """
    # Use 1.96 * std for 95% confidence interval
    margin = 1.96 * std
    lower_bound = max(0, prediction - margin)
    upper_bound = prediction + margin
    
    # Calculate confidence percentage (inverse of coefficient of variation)
    # Higher confidence when MAE is small relative to prediction
    if prediction > 0:
        cv_ratio = mae / prediction
        confidence_pct = max(0, min(100, (1 - cv_ratio) * 100))
    else:
        confidence_pct = 50.0
    
    return {
        'lower_bound': float(lower_bound),
        'upper_bound': float(upper_bound),
        'confidence_pct': float(confidence_pct)
    }


def get_feature_contributions(feature_df, demographics, weather_info, form_data, raw_feature_values=None):
    """
    Calculate and explain the top feature contributions to the prediction.
    Returns a list of feature explanations sorted by importance.
    """
    contributions = []
    
    # Check if feature importance data is available
    if FEATURE_IMPORTANCE_DF is None or len(FEATURE_IMPORTANCE_DF) == 0:
        # Fallback: use columns with recognizable names as proxies
        fallback_features = ['Duration (Hours)', 'Population', 'Median_Income', 'Industry', 'Weather_Score']
        for feature_name in fallback_features:
            if feature_name in feature_df.columns:
                feature_value = feature_df[feature_name].iloc[0]
                contributions.append({
                    'name': feature_name,
                    'importance': f"~{len(contributions)+1}",  # Relative ranking
                    'value': str(feature_value)[:50],
                    'explanation': f'Key factor in event prediction'
                })
        return contributions
    
    # Get top contributing features from the global feature importance
    top_features = FEATURE_IMPORTANCE_DF.head(10) if len(FEATURE_IMPORTANCE_DF) > 0 else pd.DataFrame()
    
    for _, row in top_features.iterrows():
        feature_name = row['feature']
        importance = row['importance']
        
        # Skip if feature not in current data
        if feature_name not in feature_df.columns:
            continue
            
        if raw_feature_values and feature_name in raw_feature_values:
            feature_value = raw_feature_values[feature_name]
        else:
            feature_value = feature_df[feature_name].iloc[0]
        
        # Convert to JSON-serializable type
        if hasattr(feature_value, 'item'):  # NumPy types
            feature_value = feature_value.item()
        elif hasattr(feature_value, 'tolist'):  # Arrays
            feature_value = feature_value.tolist()
        
        # Generate human-readable explanation
        explanation = generate_feature_explanation(
            feature_name, feature_value, importance, demographics, weather_info, form_data
        )
        
        contributions.append({
            'feature': feature_name,
            'display_name': humanize_feature_name(feature_name),
            'importance': round(importance, 1),  # importance is already in percentage
            'value': feature_value,
            'explanation': explanation
        })
    
    return contributions


def humanize_feature_name(feature_name):
    """Map model feature names to user-friendly labels."""
    if not feature_name:
        return 'Unknown Feature'

    friendly_map = {
        'Median_Income': 'Median Income',
        'Median_Household_Income': 'Median Income',
        'Total_Population': 'Population',
        'Unemployment_Rate': 'Unemployment Rate',
        'Duration (Hours)': 'Event Duration',
        'Duration': 'Event Duration',
        'Weather_Score': 'Weather Score',
        'Event_Industry': 'Industry',
        'equipment_type': 'Equipment Type',
        'Season': 'Season',
        'Time of Day': 'Time of Day',
        'TimeOfDay': 'Time of Day',
        'ZIP Cluster': 'ZIP Cluster',
        'Income_Tier': 'Income Tier',
        'Household_Size': 'Household Size'
    }

    if feature_name in friendly_map:
        return friendly_map[feature_name]

    return feature_name.replace('_', ' ').title()


def generate_feature_explanation(feature_name, value, importance, demographics, weather_info, form_data):
    """
    Generate human-readable explanations for feature contributions.
    """
    importance_pct = importance  # importance is already in percentage
    
    # Duration-based explanations
    if 'Duration' in feature_name:
        # Ensure value is numeric for duration comparisons
        try:
            duration_val = float(value)
            if duration_val > 6:
                return f"Long event duration ({duration_val:.1f} hours) significantly increases revenue potential ({importance_pct:.1f}% model weight)"
            elif duration_val < 2:
                return f"Short event duration ({duration_val:.1f} hours) may limit revenue opportunity ({importance_pct:.1f}% model weight)"
            else:
                return f"Standard event duration ({duration_val:.1f} hours) provides moderate revenue potential ({importance_pct:.1f}% model weight)"
        except (ValueError, TypeError):
            return f"Event duration feature ({value}) contributes {importance_pct:.1f}% to the prediction"
    
    # Industry-based explanations
    elif feature_name == 'Industry':
        industry_impact = {
            'Festival': 'Festivals typically generate high revenue due to broad appeal',
            'Athletics': 'Athletic events have consistent attendance patterns',
            'Community': 'Community events build loyal local attendance',
            'Education': 'Educational events attract dedicated participants',
            'Church': 'Church events benefit from community support',
            'Corporate': 'Corporate events often have higher budgets',
            'Private': 'Private events vary widely in revenue potential',
            'Daycare': 'Daycare events target specific family demographics'
        }
        impact_text = industry_impact.get(value, f"{value} events have unique market characteristics")
        return f"{impact_text} (Industry is {importance_pct:.1f}% of prediction model)"
    
    # Income and demographics
    elif 'Income' in feature_name:
        if 'Tier' in feature_name:
            tier_explanations = {
                'High': 'High-income area suggests strong spending capacity',
                'Medium-High': 'Above-average income area indicates good revenue potential',
                'Medium': 'Average income area provides moderate spending capacity',
                'Low': 'Lower income area may limit premium pricing opportunities'
            }
            explanation = tier_explanations.get(value, f"Income tier {value} affects spending patterns")
        else:
            try:
                income_val = float(value)
                if income_val > 80000:
                    explanation = f"High median income (${income_val:,.0f}) indicates strong local spending power"
                elif income_val > 60000:
                    explanation = f"Above-average median income (${income_val:,.0f}) suggests good revenue potential"
                else:
                    explanation = f"Median income (${income_val:,.0f}) may influence event pricing strategy"
            except (ValueError, TypeError):
                explanation = f"Income feature ({value}) affects local spending patterns"
        return f"{explanation} ({importance_pct:.1f}% model weight)"
    
    # Location-based explanations
    elif 'County' in feature_name or 'City' in feature_name:
        location = feature_name.split('_')[1] if '_' in feature_name else value
        return f"{location} location characteristics contribute {importance_pct:.1f}% to the prediction model"
    
    # Time-related explanations
    elif 'Time of Day' in feature_name:
        time_explanations = {
            'Morning': 'Morning events often have dedicated attendees but smaller crowds',
            'Lunchtime': 'Lunch-time events benefit from convenient scheduling',
            'Late Afternoon': 'Late afternoon timing captures after-work and school crowds',
            'Evening': 'Evening events can draw larger social crowds',
            'Night': 'Night events may have limited family attendance but strong adult participation'
        }
        time_period = feature_name.split('_')[1] if '_' in feature_name else value
        explanation = time_explanations.get(time_period, f"{time_period} timing affects attendance patterns")
        return f"{explanation} ({importance_pct:.1f}% model importance)"
    
    # Rent and housing costs
    elif 'Rent' in feature_name:
        try:
            rent_val = float(value)
            if rent_val > 1600:
                return f"High rental costs (${rent_val:,.0f}) indicate affluent area with spending capacity ({importance_pct:.1f}% model weight)"
            elif rent_val < 1000:
                return f"Lower rental costs (${rent_val:,.0f}) suggest budget-conscious community ({importance_pct:.1f}% model weight)"
            else:
                return f"Average rental costs (${rent_val:,.0f}) reflect typical local spending patterns ({importance_pct:.1f}% model weight)"
        except (ValueError, TypeError):
            return f"Rental cost feature ({value}) influences local spending patterns ({importance_pct:.1f}% model weight)"
    
    # Household characteristics
    elif 'Household' in feature_name:
        if 'Size' in feature_name:
            try:
                size_val = float(value)
                if size_val > 2.8:
                    return f"Larger households ({size_val:.1f} avg) often mean more family-oriented events ({importance_pct:.1f}% model weight)"
                else:
                    return f"Smaller households ({size_val:.1f} avg) may prefer individual or couple activities ({importance_pct:.1f}% model weight)"
            except (ValueError, TypeError):
                return f"Household size feature ({value}) influences family event appeal ({importance_pct:.1f}% model weight)"
        elif 'Children' in feature_name:
            try:
                children_val = float(value)
                return f"Family demographics ({children_val:,.0f} households with children) influence event appeal ({importance_pct:.1f}% model weight)"
            except (ValueError, TypeError):
                return f"Family demographics feature ({value}) affects event appeal ({importance_pct:.1f}% model weight)"
    
    # Season and temporal patterns
    elif feature_name == 'Season':
        season_explanations = {
            'Spring': 'Spring events benefit from renewed outdoor activity interest',
            'Summer': 'Summer events typically see highest attendance and revenue',
            'Fall': 'Fall events capture back-to-school and holiday preparation energy',
            'Winter': 'Winter events may have weather-related attendance challenges'
        }
        explanation = season_explanations.get(value, f"{value} season affects attendance patterns")
        return f"{explanation} ({importance_pct:.1f}% model importance)"
    
    # Generic explanation for other features
    else:
        return f"{feature_name} contributes {importance_pct:.1f}% to the prediction (value: {value})"


def generate_confidence_explanation(confidence_pct, cv_mae, prediction, contributions):
    """
    Generate explanation for prediction confidence level.
    
    Args:
        confidence_pct: Confidence percentage (0-100)
        cv_mae: Cross-validation Mean Absolute Error
        prediction: Predicted value
        contributions: List of feature contribution dicts with 'importance' keys, or empty list
    """
    explanations = []
    
    # Base confidence assessment
    if confidence_pct > 80:
        explanations.append("🟢 High confidence prediction - model has strong agreement on similar events")
    elif confidence_pct > 60:
        explanations.append("🟡 Moderate confidence - model shows reasonable consistency for this type of event")
    else:
        explanations.append("🟠 Lower confidence - event characteristics show more variability in historical data")
    
    # MAE-based explanation
    mae_explanation = f"Historical model accuracy: ±${cv_mae:.0f} per hour on similar events"
    explanations.append(mae_explanation)
    
    # Feature stability assessment (only if contributions provided)
    if contributions and isinstance(contributions, list) and len(contributions) > 0:
        top_contributions = [c for c in contributions if isinstance(c, dict) and c.get('importance', 0) > 5.0]
        if len(top_contributions) >= 3:
            explanations.append("✅ Prediction based on multiple strong indicators, increasing reliability")
        else:
            explanations.append("⚠️ Prediction relies on fewer strong indicators, consider as estimate")
    else:
        explanations.append("ℹ️ Real-time prediction using merged ensemble model")
    
    # Prediction magnitude assessment
    if prediction > 200:
        explanations.append("💰 High revenue prediction - validate with local market conditions")
    elif prediction < 50:
        explanations.append("📊 Conservative revenue estimate - consider cost management")
    
    return explanations


def prepare_features(form_data):
    """
    Prepare feature dictionary from form input using the actual preprocessing pipeline.
    Fetches demographics from Census API and weather data, then appends to historical data for proper Phase 2 features.
    """
    # Extract user inputs
    industry = form_data.get('industry', 'Community')
    city = form_data.get('city', 'Naperville')
    zip_code = form_data.get('zip_code', '60540')
    event_date = form_data.get('event_date')
    start_time = form_data.get('start_time', '09:00')
    duration = float(form_data.get('duration', 4.0))
    
    # Get optional weather inputs from user
    user_temperature = form_data.get('temperature')  # Optional temperature input
    
    # Fetch demographics from Census API
    print(f"Fetching demographics for ZIP {zip_code}...")
    demographics = get_demographics_from_zip(zip_code)
    print(f"Demographics source: {demographics['source']}")
    
    # Get county from ZIP
    county = get_county_from_zip(zip_code)
    
    # Combine date and time
    start_datetime_str = f"{event_date} {start_time}"
    start_dt = datetime.strptime(start_datetime_str, '%Y-%m-%d %H:%M')
    end_dt = start_dt + pd.Timedelta(hours=duration)
    
    # Get weather data for the event
    print(f"Fetching weather data for event...")
    weather_info = get_weather_for_event(zip_code, start_dt, user_temperature)
    print(f"Weather data: {weather_info['temperature_f']}°F, rain={weather_info['is_raining']}, source={weather_info['source']}")
    
    # Format datetimes to match training CSV exactly
    start_str_iso = start_dt.strftime('%Y-%m-%d %H:%M:%S')
    end_str_formatted = end_dt.strftime('%b %d, %Y %I:%M %p')
    
    # Calculate placeholder net sales
    net_sales_placeholder = 200.0
    net_sales = net_sales_placeholder * duration
    
    # Create CSV row matching training format (include weather if available in training data)
    new_event = {
        'Industry': industry,
        'City': city,
        'County': county,
        'Zip Code': int(zip_code),
        'Event Start Date Time': start_str_iso,
        'Event End Date Time': end_str_formatted,
        'Net Sales': net_sales,
        'Population': demographics['Total_Population'],
        'Population_Density': demographics['Total_Population'] / 100,
        'Median_Income': demographics['Median_Household_Income'],
        'Avg_Household_Size': demographics['Average_Household_Size'],
        'Unemployment_Rate': demographics['Unemployment_Rate'],
        'Higher_Education_Rate': 40.0,
        'Households_With_Children': demographics['Households_With_Children'],
        'Median_Rent': demographics['Median_Rent'],
        'Median_Home_Value': demographics['Median_Home_Value'],
        'Affordability_Ratio': demographics['Median_Home_Value'] / demographics['Median_Household_Income'],
        'Rent_To_Income_Ratio': demographics['Rent_To_Income_Ratio'],
        'Demographics_Is_Estimated': demographics['source'] == 'default',
        # Add weather data
        'Temperature_F': weather_info['temperature_f'],
        'Is_Raining': weather_info['is_raining']
    }
    
    # Append to historical data and reprocess to get Phase 2 features
    temp_df = pd.DataFrame([new_event])
    temp_csv = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='')
    temp_df.to_csv(temp_csv.name, index=False)
    temp_csv.close()
    
    try:
        # Process the new event through pipeline
        # print(f"Processing new event through load_and_preprocess...")
        # new_event_processed = load_and_preprocess(temp_csv.name)
        # NOTE: This function uses old prepare_features logic and is deprecated
        # Use prepare_features_v2 instead
        raise NotImplementedError("Old prepare_features path - use prepare_features_v2 instead")
        
        print(f"New event processed shape: {new_event_processed.shape}")
        print(f"New event columns: {list(new_event_processed.columns)[:20]}...")
        
        # Use the same drop columns as training
        drop_cols = TRAINING_DROP_COLS
        feature_df = new_event_processed.drop(columns=[col for col in drop_cols if col in new_event_processed.columns])
        
        print(f"Feature DataFrame shape after dropping metadata: {feature_df.shape}")
        print(f"Expected: {len(FEATURE_COLUMNS)} features")
        
        # Verify we have the right number of features
        if feature_df.shape[1] != len(FEATURE_COLUMNS):
            print(f"WARNING: Feature count mismatch! Got {feature_df.shape[1]}, expected {len(FEATURE_COLUMNS)}")
            print(f"Missing columns might be Phase 2 features that need historical context")
            
            # Get missing columns by comparing with historical data feature columns
            current_cols = set(feature_df.columns)
            expected_cols = set(FEATURE_COLUMNS)
            missing = expected_cols - current_cols
            extra = current_cols - expected_cols
            
            if extra:
                print(f"Extra {len(extra)} columns (will be dropped): {list(extra)[:10]}...")
                # Drop extra columns
                feature_df = feature_df.drop(columns=list(extra))
            
            if missing:
                print(f"Missing {len(missing)} columns: {list(missing)[:10]}...")
                # Add missing columns with default values
                for col in missing:
                    if col in HISTORICAL_DATA.columns:
                        # Use median from historical data as default
                        if HISTORICAL_DATA[col].dtype in ['float64', 'int64']:
                            feature_df[col] = HISTORICAL_DATA[col].median()
                        else:
                            feature_df[col] = HISTORICAL_DATA[col].mode()[0] if len(HISTORICAL_DATA[col].mode()) > 0 else 'Unknown'
                
                # Reorder columns to match historical data
                feature_df = feature_df[FEATURE_COLUMNS]
                print(f"After adding missing columns and reordering: {feature_df.shape}")
        
        # Convert all category dtypes to object/string (CatBoost doesn't like pandas category dtype)
        for col in feature_df.columns:
            if feature_df[col].dtype.name == 'category':
                feature_df[col] = feature_df[col].astype(str)
        
        # Convert categorical columns to string type to match training
        for cat_col in CATEGORICAL_FEATURES:
            if cat_col in feature_df.columns:
                feature_df[cat_col] = feature_df[cat_col].astype(str)
        
        # Debug: print column order and dtypes
        print(f"\nFinal feature columns ({len(feature_df.columns)}):")
        datetime_cols = []
        object_cols = []
        numeric_cols = []
        other_cols = []
        
        for i, col in enumerate(feature_df.columns):
            dtype = feature_df[col].dtype
            val = feature_df[col].iloc[0]
            cat_marker = " [CAT]" if col in CATEGORICAL_FEATURES else ""
            
            # Categorize by dtype
            if pd.api.types.is_datetime64_any_dtype(dtype):
                datetime_cols.append(col)
                print(f"  {i}: {col} ({dtype}){cat_marker} = {val} *** DATETIME - SHOULD BE DROPPED ***")
            elif dtype == 'object':
                object_cols.append(col)
                print(f"  {i}: {col} ({dtype}){cat_marker} = {val}")
            elif pd.api.types.is_numeric_dtype(dtype):
                numeric_cols.append(col)
                if i < 20 or col in CATEGORICAL_FEATURES:  # Only print first 20 numeric or if marked as categorical
                    print(f"  {i}: {col} ({dtype}){cat_marker} = {val}")
            else:
                other_cols.append(col)
                print(f"  {i}: {col} ({dtype}){cat_marker} = {val} *** UNKNOWN TYPE ***")
        
        print(f"\nSummary: {len(numeric_cols)} numeric, {len(object_cols)} object, {len(datetime_cols)} datetime, {len(other_cols)} other")
        
        if datetime_cols:
            print(f"\nERROR: Found datetime columns that should have been dropped: {datetime_cols}")
            print("Dropping them now...")
            feature_df = feature_df.drop(columns=datetime_cols)
            print(f"New shape after dropping datetime cols: {feature_df.shape}")
        
        return feature_df, demographics, weather_info
    except Exception as e:
        print(f"Error in prepare_features: {str(e)}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        # Clean up temp file
        if os.path.exists(temp_csv.name):
            os.unlink(temp_csv.name)


def prepare_features_v2(form_data):
    """
    Create features matching the complete 71-feature engineered dataset.
    Uses the same Phase 1 + Phase 2 feature engineering as training.
    """
    print("=== PREPARING FEATURES (COMPLETE 71-FEATURE SET) ===")
    print(f"Form data keys: {list(form_data.keys())}")
    
    try:
        # Extract form data
        industry = form_data.get('industry')
        city = form_data.get('city', 'Tinley Park')
        zip_code = int(form_data.get('zip_code', 60487))
        duration = float(form_data.get('duration', 4.0))
        
        # Parse date and time
        event_date = form_data.get('event_date')
        start_time = form_data.get('start_time', '10:00')
        start_datetime_str = f"{event_date} {start_time}"
        start_datetime = datetime.strptime(start_datetime_str, '%Y-%m-%d %H:%M')
        end_datetime = start_datetime + pd.Timedelta(hours=duration)
        
        # Get demographics
        demographics = get_demographics_from_zip(zip_code)
        
        # Get weather data  
        weather_info = get_weather_for_event(zip_code, start_datetime, 
                                           user_temperature=form_data.get('temperature'))
        
        print(f"Event: {industry} in {city} on {event_date}")
        print(f"Demographics source: {demographics['source']}")
        print(f"Weather data: {weather_info['temperature_f']:.1f}°F, rain={weather_info['is_raining']}, source={weather_info['source']}")
        
        # Create raw event data with ALL required base columns
        event_data = {
            'Industry': industry,
            'City': city,
            'County': demographics.get('County', 'Cook County'),
            'Zip Code': zip_code,
            'Event Start Date Time': start_datetime,
            'Event End Date Time': end_datetime,
            'Event Start Date': start_datetime,  # Phase 1 requires this
            'Event End Date': end_datetime,      # For consistency
            'Net Sales': 0.0,  # Placeholder, not used in prediction
            'Duration (Hours)': duration,
            'Population': demographics['Total_Population'],
            'Population_Density': demographics.get('Population_Density', demographics['Total_Population'] / 100),
            'Median_Income': demographics['Median_Household_Income'],
            'Avg_Household_Size': demographics['Average_Household_Size'],
            'Unemployment_Rate': demographics['Unemployment_Rate'],
            'Higher_Education_Rate': 40.0,
            'Households_With_Children': demographics['Households_With_Children'],
            'Median_Rent': demographics['Median_Rent'],
            'Median_Home_Value': demographics['Median_Home_Value'],
            'Affordability_Ratio': demographics['Median_Home_Value'] / demographics['Median_Household_Income'],
            'Rent_To_Income_Ratio': demographics['Rent_To_Income_Ratio'],
            'Temperature_F': weather_info['temperature_f'],
            'Is_Raining': int(weather_info['is_raining']),
            'Net Sales per Hour': 0.0,  # Placeholder
        }
        
        # Create DataFrame
        df = pd.DataFrame([event_data])
        
        # Add basic temporal features (done before Phase 1 in training pipeline)
        df['Month'] = df['Event Start Date'].dt.month
        df['Day of Week'] = df['Event Start Date'].dt.dayofweek
        df['Is Weekend'] = df['Event Start Date'].dt.dayofweek.isin([5, 6]).astype(int)
        df['Week of Month'] = df['Event Start Date'].dt.day.apply(lambda x: (x-1)//7 + 1)
        df['Event Start Hour'] = df['Event Start Date'].dt.hour
        
        # Season
        df['Season'] = df['Event Start Date'].dt.month.map({
            12: 'Winter', 1: 'Winter', 2: 'Winter',
            3: 'Spring', 4: 'Spring', 5: 'Spring',
            6: 'Summer', 7: 'Summer', 8: 'Summer',
            9: 'Fall', 10: 'Fall', 11: 'Fall'
        })
        
        # Time of Day (based on start hour + half duration)
        def get_time_of_day(hour):
            if 6 <= hour < 10:
                return 'Morning'
            elif 10 <= hour < 14:
                return 'Lunchtime'
            elif 14 <= hour < 17:
                return 'Early Afternoon'
            elif 17 <= hour < 19:
                return 'Late Afternoon'
            else:
                return 'Night'
        
        df['Time of Day'] = (df['Event Start Hour'] + (df['Duration (Hours)'] / 2)).apply(get_time_of_day)
        
        # Industry Category
        industry_categories = {
            'Community': 'Community/Public',
            'Church': 'Religious',
            'Private': 'Private Events',
            'Athletics': 'Sports/Recreation',
            'Education': 'Education',
            'Corporate': 'Business',
            'Festival': 'Entertainment',
            'Daycare': 'Childcare'
        }
        df['Industry Category'] = df['Industry'].map(industry_categories).fillna('Other')
        
        # ZIP Cluster (first 3 digits)
        zip_cluster = predict_zip_cluster(df['Zip Code'].iloc[0])
        try:
            df['ZIP Cluster'] = int(zip_cluster)
        except (TypeError, ValueError):
            df['ZIP Cluster'] = 0
        
        # Events in County (placeholder for single event)
        df['Events in County'] = 1
        
        # Income and Population Tiers
        try:
            df['Income_Tier'] = pd.qcut([df['Median_Income'].iloc[0]], q=4, labels=['Low', 'Medium', 'High', 'Very High'], duplicates='drop')[0]
        except:
            df['Income_Tier'] = 'Medium'  # Default
        
        try:
            df['Population_Tier'] = pd.qcut([df['Population_Density'].iloc[0]], q=4, labels=['Low', 'Medium', 'High', 'Very High'], duplicates='drop')[0]
        except:
            df['Population_Tier'] = 'Medium'  # Default
        
        # Apply Phase 1 and Phase 2 feature engineering (just like training)
        try:
            from Kona_AI_ML.outdated.feature_engineering import apply_phase1_feature_engineering
        except:
            from Kona_AI_ML.outdated.feature_engineering import apply_phase1_feature_engineering
        
        # Debug: Check columns before Phase 1
        print(f"Columns BEFORE Phase 1: {len(df.columns)}")
        # Apply Phase 1 feature engineering (temporal, cyclical, interactions, composite scores)
        df = apply_phase1_feature_engineering(df)
        # Debug: Check columns after Phase 1
        print(f"Columns AFTER Phase 1: {len(df.columns)}")
        
        print(f"All columns after Phase 1: {list(df.columns)}")
        
        
        # For real-time predictions, we don't have historical context to compute Phase 2 features
        # Instead, use median values from training data as reasonable defaults
        # (In production, each franchise will have their own historical data and pre-computed features)
        PHASE2_DEFAULTS = {
            'Industry_Avg_7D': 156.84,
            'Industry_Avg_14D': 157.01,
            'Industry_Avg_30D': 154.93,
            'Same_Week_Industry_Avg': 159.91,
            'County_Avg_30D': 174.90,
            'Events_Past_30d': 11.0,
            'Competing_Events_Same_Day': 1.0,
            'Competing_Events_Same_Industry': 0.0,
            'Industry_Share_In_County': 0.47,
            'Industry_Growth_Rate': -0.15,
            'County_Momentum': -0.66,
            'Industry_Seasonality': 0.21
        }
        
        # Add Phase 2 features with default values
        for col, default_val in PHASE2_DEFAULTS.items():
            df[col] = default_val
        
        # Drop columns that shouldn't be used for prediction (MUST match training exactly!)
        # Use the same list as defined at the top to ensure consistency
        drop_cols = TRAINING_DROP_COLS  # This includes Zip Code, City, County
        print(f"DROP COLS: {drop_cols}")
        print(f"Before drop: {len(df.columns)} columns, has pricing: Is_Free={'Is_Free' in df.columns}")
        df = df.drop(columns=[col for col in drop_cols if col in df.columns])
        print(f"After drop: {len(df.columns)} columns, has pricing: Is_Free={'Is_Free' in df.columns}")
        
        print(f"After Phase 1+2 and dropping: {df.shape[1]} features")
        
        # Ensure columns are in the EXACT order as training data
        # Load the expected column order AND dtypes from the complete features dataset
        training_sample = pd.read_csv(TRAINING_DATA_PATH, nrows=1)
        # Remove the columns we dropped
        for col in drop_cols:
            if col in training_sample.columns:
                training_sample = training_sample.drop(columns=[col])
        
        expected_order = training_sample.columns.tolist()
        expected_dtypes = training_sample.dtypes.to_dict()
        
        # Check for genuinely missing columns and fail early
        missing_cols = set(expected_order) - set(df.columns)
        if missing_cols:
            print(f"\n[DEBUG] Expected columns ({len(expected_order)}): {expected_order[:10]}... (truncated)")
            print(f"[DEBUG] DataFrame columns ({len(df.columns)}): {list(df.columns)[:10]}... (truncated)")
            print(f"[DEBUG] Missing {len(missing_cols)} columns:")
            for mc in list(missing_cols)[:15]:
                print(f"  - {mc}")
            raise ValueError(f"Missing {len(missing_cols)} required features: {list(missing_cols)[:10]}")
        
        # Reorder columns to match training exactly
        df = df[expected_order]
        
        # Match dtypes EXACTLY to training data
        for col in df.columns:
            if col in expected_dtypes:
                expected_dtype = expected_dtypes[col]
                if expected_dtype == 'object':
                    # Convert to category for XGBoost
                    df[col] = df[col].astype('category')
                elif str(df[col].dtype) != str(expected_dtype):
                    # Convert numeric dtypes to match exactly
                    df[col] = df[col].astype(expected_dtype)
        
        print(f"Final feature DataFrame shape: {df.shape}")
        print(f"Dtypes match training: first feature dtype={df[df.columns[0]].dtype}, 24th={df[df.columns[23]].dtype}")
        
        return df
        
    except Exception as e:
        print(f"Feature creation error: {e}")
        import traceback
        traceback.print_exc()
        return None


def prepare_features_wrapper(form_data):
    """
    Wrapper to provide interface for feature preparation.
    Uses prepare_features_v2 as the main implementation.
    Returns: feature_df, demographics, weather_info
    """
    try:
        # Use prepare_features_v2 which is the main implementation
        feature_df = prepare_features_v2(form_data)
        
        if feature_df is None:
            return None, None, None
        
        # Get demographics and weather separately for display
        zip_code = str(form_data.get('zip_code', '60487')).zfill(5)
        event_date = form_data.get('event_date')
        start_time = form_data.get('start_time', '10:00')
        
        try:
            start_datetime_str = f"{event_date} {start_time}"
            start_datetime = datetime.strptime(start_datetime_str, '%Y-%m-%d %H:%M')
        except (ValueError, TypeError):
            start_datetime = datetime.now()
        
        demographics = get_demographics_from_zip(zip_code)
        weather_info = get_weather_for_event(zip_code, start_datetime, 
                                           user_temperature=form_data.get('temperature'))
        
        return feature_df, demographics, weather_info
        
    except Exception as e:
        print(f"Wrapper error: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None


# ===== AUTHENTICATION ROUTES =====

@app.route('/register', methods=['GET', 'POST'])
def register_page():
    """Display registration page or create a new franchise account."""
    if request.method == 'GET':
        return render_template('register.html')

    franchise_id = request.form.get('franchise_id', '').strip()
    franchise_name = request.form.get('franchise_name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    target_str = request.form.get('target_net_sales_per_hour', '').strip()

    if not franchise_id or not franchise_name or not email or not password:
        return render_template('register.html', error='All fields except target are required')

    if len(password) < 8:
        return render_template('register.html', error='Password must be at least 8 characters')

    target_value = None
    if target_str:
        try:
            target_value = float(target_str)
            if target_value <= 0:
                return render_template('register.html', error='Target must be a positive number')
        except ValueError:
            return render_template('register.html', error='Target must be a number')

    if franchise_db.get_franchise(franchise_id):
        return render_template('register.html', error='Franchise ID already exists')
    if franchise_db.get_franchise_by_email(email):
        return render_template('register.html', error='Email already registered')

    password_hash = auth_manager.hash_password(password)
    created = franchise_db.create_franchise(
        franchise_id=franchise_id,
        franchise_name=franchise_name,
        email=email,
        password_hash=password_hash,
        target_net_sales_per_hour=target_value
    )

    if not created:
        return render_template('register.html', error='Account already exists or could not be created')

    # Auto-login after registration
    token, error = auth_manager.login(franchise_id, password)
    if error or not token:
        return redirect(url_for('login_page'))

    response = make_response(redirect(url_for('dashboard')))
    response.set_cookie(
        SessionManager.get_session_cookie_name(),
        token,
        max_age=SessionManager.get_session_duration(),
        httponly=True,
        samesite='Lax'
    )

    return response

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """Display login page or process login request."""
    if request.method == 'GET':
        return render_template('login.html')
    
    # Handle login POST request
    franchise_id = request.form.get('franchise_id', '').strip()
    password = request.form.get('password', '').strip()
    
    if not franchise_id or not password:
        return render_template('login.html', error='Please provide franchise ID and password')
    
    # Attempt login
    token, error = auth_manager.login(franchise_id, password)
    
    if error:
        return render_template('login.html', error=error)
    
    # Create response and set session cookie
    next_page_raw = request.args.get('next', '')
    next_page_candidate = next_page_raw.replace('\\', '')
    parsed_next = urlparse(next_page_candidate)
    if (
        next_page_candidate
        and not parsed_next.scheme
        and not parsed_next.netloc
        and next_page_candidate.startswith('/')
    ):
        next_page = next_page_candidate
    else:
        next_page = url_for('dashboard')
    response = make_response(redirect(next_page))
    response.set_cookie(
        SessionManager.get_session_cookie_name(),
        token,
        max_age=SessionManager.get_session_duration(),
        httponly=True,
        samesite='Lax'
    )
    
    return response


@app.route('/logout')
def logout():
    """Logout the current franchise."""
    session_token = request.cookies.get(SessionManager.get_session_cookie_name())
    if session_token:
        auth_manager.logout(session_token)
    
    response = make_response(redirect(url_for('login_page')))
    response.delete_cookie(SessionManager.get_session_cookie_name())
    return response


@app.route('/dashboard')
@require_auth
def dashboard():
    """Display franchise dashboard."""
    franchise = get_current_franchise()
    threshold = PROD_MANAGER.FRANCHISE_DATA_THRESHOLD if PROD_MANAGER else 100
    franchise_progress = franchise_db.get_franchise_model_progress(franchise['franchise_id'], threshold=threshold)
    franchise_training_result = None
    should_train_initial = franchise_progress['ready_for_training'] and not franchise_progress['model_exists']
    should_retrain = franchise_progress.get('should_retrain_now', False)
    if FRANCHISE_MODEL_TRAINER and (should_train_initial or should_retrain):
        franchise_training_result = FRANCHISE_MODEL_TRAINER.ensure_model(
            franchise['franchise_id'],
            force=should_retrain,
        )
        franchise_progress = franchise_training_result.get('progress') or franchise_db.get_franchise_model_progress(
            franchise['franchise_id'],
            threshold=threshold,
        )

    available_models = get_franchise_available_models(franchise['franchise_id'], franchise_progress)

    dashboard_state = _get_dashboard_state(request.args)
    month_window = _dashboard_month_bounds(dashboard_state['month'])
    dashboard_stats = franchise_db.get_prediction_dashboard_stats(
        franchise['franchise_id'],
        month_start=month_window['start'],
        month_end_exclusive=month_window['next_start'],
    )

    # Get recent predictions from database
    recent_predictions, total_predictions = franchise_db.get_recent_predictions_page(
        franchise['franchise_id'],
        page=dashboard_state['page'],
        page_size=DASHBOARD_PAGE_SIZE,
        status_filter=dashboard_state['status'],
        sort_by=dashboard_state['sort_by'],
        sort_dir=dashboard_state['sort_dir'],
        month_start=month_window['start'],
        month_end_exclusive=month_window['next_start'],
    )

    total_pages = max(1, (total_predictions + DASHBOARD_PAGE_SIZE - 1) // DASHBOARD_PAGE_SIZE)
    if dashboard_state['page'] > total_pages and total_predictions > 0:
        return redirect(url_for(
            'dashboard',
            page=total_pages,
            month=dashboard_state['month'],
            status=dashboard_state['status'],
            sort_by=dashboard_state['sort_by'],
            sort_dir=dashboard_state['sort_dir'],
        ))

    has_test_predictions = any(pred.get('is_test') for pred in recent_predictions)
    for pred in recent_predictions:
        pred['event_status'] = _normalize_event_status(pred.get('event_status'))
        actual_value = pred.get('actual_total_net_sales')
        predicted_total = pred.get('predicted_total_revenue')
        if actual_value is None or predicted_total is None:
            pred['predicted_vs_actual_diff'] = None
        else:
            pred['predicted_vs_actual_diff'] = float(predicted_total) - float(actual_value)
    
    # Try to get default model from database, fallback to first available
    default_model_id = franchise.get('default_model_id')
    default_model = None
    
    if default_model_id:
        default_model = next((m for m in available_models if m['model_id'] == default_model_id), None)
    
    if not default_model and available_models:
        default_model = next((model for model in available_models if model.get('selectable', True)), available_models[0])

    franchise_model_popup = None
    popup_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if franchise_progress['ready_for_training'] and not franchise_progress.get('threshold_popup_shown_at'):
        if franchise_training_result and franchise_training_result.get('status') == 'awaiting_feature_snapshots':
            popup_message = franchise_training_result['message']
        else:
            popup_message = 'You hit the threshold for a dedicated franchise model. We have started preparing it, and it should be ready by your next dashboard visit.'
        franchise_model_popup = {
            'kind': 'threshold',
            'title': 'Franchise Model Unlocked',
            'message': popup_message,
        }
        franchise_db.update_franchise_model_status(
            franchise['franchise_id'],
            threshold_popup_shown_at=popup_timestamp,
        )
    elif franchise_progress['model_exists'] and franchise_progress.get('threshold_popup_shown_at') and not franchise_progress.get('model_ready_popup_shown_at'):
        ready_model = franchise_progress.get('latest_model') or {}
        ready_metadata = _parse_training_metadata(ready_model.get('training_metadata_json'))
        ready_metrics = ready_metadata.get('metrics', {}) if isinstance(ready_metadata, dict) else {}
        ready_r2 = ready_metrics.get('r2')
        if ready_r2 is None:
            ready_r2 = ready_model.get('r2')
        franchise_model_popup = {
            'kind': 'ready',
            'title': 'Your Franchise Model Is Ready',
            'message': 'Your dedicated model is now available in Model Selection with its validation stats.',
            'cv_mae': ready_model.get('cv_mae'),
            'r2': ready_r2,
            'training_rows': ready_model.get('data_records_count'),
        }
        franchise_db.update_franchise_model_status(
            franchise['franchise_id'],
            model_ready_popup_shown_at=popup_timestamp,
        )
    elif franchise_training_result and franchise_training_result.get('status') == 'retrained' and not franchise_progress.get('retrain_popup_shown_at'):
        franchise_model_popup = {
            'kind': 'retrained',
            'title': 'Franchise Model Refreshed',
            'message': (
                f"Your model was automatically refreshed using {franchise_progress.get('last_trained_event_count', 0)} completed events. "
                f"Next refresh target: {franchise_progress.get('next_retrain_event_count', 'TBD')} completed events."
            ),
        }
        franchise_db.update_franchise_model_status(
            franchise['franchise_id'],
            retrain_popup_shown_at=popup_timestamp,
        )
    
    return render_template('dashboard.html',
                          franchise_name=franchise['franchise_name'],
                          models=available_models,
                          default_model=default_model,
                          franchise_model_progress=franchise_progress,
                          franchise_model_popup=franchise_model_popup,
                          recent_predictions=recent_predictions,
                          realized_count=dashboard_stats['realized_count'],
                          realized_total=dashboard_stats['realized_total'],
                          completed_predicted_total=dashboard_stats['completed_predicted_total'],
                          realized_vs_predicted_delta=dashboard_stats['realized_vs_predicted_delta'],
                          realized_vs_predicted_pct=dashboard_stats['realized_vs_predicted_pct'],
                          monthly_actual_net_sales_per_hour=dashboard_stats['monthly_actual_net_sales_per_hour'],
                          forecast_count=dashboard_stats['forecast_count'],
                          booked_count=dashboard_stats['booked_count'],
                          needs_outcome_count=dashboard_stats['needs_outcome_count'],
                          target_net_sales_per_hour=franchise.get('target_net_sales_per_hour'),
                          has_test_predictions=has_test_predictions,
                          page=dashboard_state['page'],
                          month_key=dashboard_state['month'],
                          month_label=month_window['label'],
                          prev_month_key=month_window['prev_key'],
                          next_month_key=month_window['next_key'],
                          page_size=DASHBOARD_PAGE_SIZE,
                          total_pages=total_pages,
                          total_predictions=total_predictions,
                          status_filter=dashboard_state['status'],
                          sort_by=dashboard_state['sort_by'],
                          sort_dir=dashboard_state['sort_dir'])


@app.route('/predictions/update-actual', methods=['POST'])
@require_auth
def update_actual_net_sales():
    """Update actual net sales for a prediction."""
    prediction_id = request.form.get('prediction_id', '').strip()
    actual_str = request.form.get('actual_total_net_sales', '').strip()

    dashboard_state = _get_dashboard_state(request.form)

    if not prediction_id or not actual_str:
        return redirect(url_for('dashboard', **dashboard_state))

    try:
        actual_total = float(actual_str)
    except ValueError:
        return redirect(url_for('dashboard', **dashboard_state))

    if actual_total < 0:
        return redirect(url_for('dashboard', **dashboard_state))

    franchise_db.update_actual_net_sales(
        get_current_franchise_id(),
        prediction_id,
        actual_total
    )

    return redirect(url_for('dashboard', **dashboard_state))


@app.route('/predictions/update-status', methods=['POST'])
@require_auth
def update_prediction_status():
    """Update lifecycle status for a prediction."""
    prediction_id = request.form.get('prediction_id', '').strip()
    event_status = _normalize_event_status(request.form.get('event_status'))

    dashboard_state = _get_dashboard_state(request.form)

    if not prediction_id:
        return redirect(url_for('dashboard', **dashboard_state))

    franchise_db.update_prediction_status(
        get_current_franchise_id(),
        prediction_id,
        event_status,
        include_in_training=False,
    )
    return redirect(url_for('dashboard', **dashboard_state))


@app.route('/predictions/delete-tests', methods=['POST'])
@require_auth
def delete_test_predictions():
    """Delete selected test predictions for the current franchise."""
    payload = request.get_json(silent=True) or {}
    prediction_ids = payload.get('prediction_ids') or []
    if not isinstance(prediction_ids, list):
        return jsonify({'success': False, 'error': 'Invalid prediction list'}), 400

    deleted = franchise_db.delete_predictions(
        get_current_franchise_id(),
        [str(pid) for pid in prediction_ids],
        test_only=True
    )

    return jsonify({'success': True, 'deleted': deleted})


@app.route('/select-model', methods=['POST'])
@require_auth
def select_model():
    """Update the franchise's default model."""
    franchise = get_current_franchise()
    model_id = request.form.get('model_id', '').strip()
    month_key = _normalize_dashboard_month(request.form.get('month'))
    
    if not model_id:
        return redirect(url_for('dashboard', month=month_key))
    
    # Verify model is actually available
    available_models = get_franchise_available_models(franchise['franchise_id'])
    selected_model = next((m for m in available_models if m['model_id'] == model_id), None)
    if not selected_model or not selected_model.get('selectable', True):
        return redirect(url_for('dashboard', month=month_key))
    
    # Update default model in database
    franchise_db.update_franchise_default_model(franchise['franchise_id'], model_id)
    
    return redirect(url_for('dashboard', month=month_key))


@app.route('/update-target', methods=['POST'])
@require_auth
def update_target():
    """Update franchise target net sales per hour."""
    franchise = get_current_franchise()
    target_value = request.form.get('target_net_sales_per_hour', '').strip()
    month_key = _normalize_dashboard_month(request.form.get('month'))

    try:
        target_float = float(target_value)
    except (TypeError, ValueError):
        return redirect(url_for('dashboard', month=month_key))

    if target_float <= 0:
        return redirect(url_for('dashboard', month=month_key))

    franchise_db.update_franchise_target(franchise['franchise_id'], target_float)
    return redirect(url_for('dashboard', month=month_key))


@app.route('/')
def home():
    """Redirect to login if not authenticated, otherwise show prediction form."""
    session_token = request.cookies.get(SessionManager.get_session_cookie_name())
    if not session_token:
        return redirect(url_for('login_page'))
    
    # Validate session
    franchise_id, error = auth_manager.validate_session(session_token)
    if error:
        return redirect(url_for('login_page'))
    
    # Render the main prediction form
    return render_template('index.html')


@app.route('/predict', methods=['POST'])
def predict():
    """Handle prediction requests using clean_layers production model."""
    # Check authentication but return JSON error if not authenticated
    session_token = request.cookies.get(SessionManager.get_session_cookie_name())
    if not session_token:
        return jsonify({
            'success': False,
            'error': 'Authentication required. Please log in.',
            'redirect': url_for('login_page')
        }), 401
    
    franchise_id, error = auth_manager.validate_session(session_token)
    if error or not franchise_id:
        return jsonify({
            'success': False,
            'error': 'Session expired. Please log in again.',
            'redirect': url_for('login_page')
        }), 401
    
    try:
        # Import clean_layers feature preparation
        from clean_layers_feature_preparation import prepare_features_for_clean_layers_model
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Infer has_smoothie_capability from equipment_type (only blended truck has it)
        form_data['has_smoothie_capability'] = '1' if form_data.get('equipment_type') == 'blended truck' else '0'
        
        # Get demographics and weather
        zip_code = str(form_data.get('zip_code', '60487')).zfill(5)
        event_date_str = form_data.get('event_date', datetime.now().strftime('%Y-%m-%d'))
        start_time_str = form_data.get('start_time', '18:00')
        
        demographics = get_demographics_from_zip(zip_code)
        
        try:
            event_date = datetime.strptime(event_date_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            event_date = datetime.now()
        
        weather_info = get_weather_for_event(zip_code, event_date, 
                                            user_temperature=form_data.get('temperature'))
        
        # Prepare exactly 46 features for clean_layers model
        feature_df, raw_feature_values = prepare_features_for_clean_layers_model(
            form_data=form_data,
            franchise_id=franchise_id,
            demographics_dict=demographics,
            weather_dict=weather_info,
            franchise_history=None,  # TODO: Load from database if available
            return_raw_values=True
        )
        
        if feature_df is None or feature_df.empty:
            return jsonify({'error': 'Failed to prepare features'}), 400
        
        # Make prediction using ProductionModelManager
        if PROD_MANAGER is None:
            return jsonify({'error': 'Model not loaded. Please try again.'}), 500
        
        predictions, model_type = PROD_MANAGER.predict(feature_df, franchise_id=franchise_id)
        y_hat = float(predictions[0]) if hasattr(predictions, '__len__') else float(predictions)
        
        # Get prediction from model
        # Model predicts Net_Sales (total sales for the event)
        duration = float(form_data.get('duration', 4.0))
        total_net_sales = y_hat  # y_hat is already the total Net_Sales prediction
        net_sales_per_hour = total_net_sales / duration if duration > 0 else total_net_sales
        
        # Confidence intervals (based on MAE)
        total_lower = max(0.0, total_net_sales * 0.8)  # Conservative 80%
        total_upper = total_net_sales * 1.2  # Optimistic 120%
        per_hour_lower = total_lower / duration if duration > 0 else total_lower
        per_hour_upper = total_upper / duration if duration > 0 else total_upper
        confidence_pct = 80.0
        
        # Generate explanations (feature contributions + advanced explanations if available)
        feature_contributions = get_feature_contributions(
            feature_df,
            demographics,
            weather_info,
            form_data,
            raw_feature_values=raw_feature_values
        )

        advanced_explanation = None
        prediction_explainer = get_prediction_explainer()
        if prediction_explainer is not None:
            try:
                advanced_explanation = prediction_explainer.explain_prediction(
                    feature_df,
                    y_hat,
                    top_k=10
                )
            except Exception as e:
                print(f"[WARN] Advanced explanation failed: {e}")

        # If global importances are missing, fall back to advanced feature impacts
        if not feature_contributions and advanced_explanation:
            adv_features = advanced_explanation.get('feature_contributions', [])
            for contrib in adv_features[:6]:
                impact_pct = float(contrib.get('importance_pct', 0.0))
                impact_val = float(contrib.get('impact', 0.0))
                direction = str(contrib.get('direction', 'affects')).capitalize()
                feature_contributions.append({
                    'feature': contrib.get('feature', 'Unknown'),
                    'display_name': humanize_feature_name(contrib.get('feature', 'Unknown')),
                    'importance': round(impact_pct, 1),
                    'value': contrib.get('value', ''),
                    'explanation': f"{direction} prediction (impact: ${impact_val:,.2f})"
                })

        if advanced_explanation:
            adv_contribs = advanced_explanation.get('feature_contributions', [])
            for contrib in adv_contribs:
                feature_name = contrib.get('feature', '')
                contrib['display_name'] = humanize_feature_name(feature_name)
                if raw_feature_values and feature_name in raw_feature_values:
                    contrib['value'] = raw_feature_values[feature_name]

            shap_values = advanced_explanation.get('shap_values', {})
            shap_features = shap_values.get('top_features', []) if isinstance(shap_values, dict) else []
            for feat in shap_features:
                feature_name = feat.get('feature', '')
                feat['display_name'] = humanize_feature_name(feature_name)
                if raw_feature_values and feature_name in raw_feature_values:
                    feat['value'] = raw_feature_values[feature_name]
        
        confidence_explanations = generate_confidence_explanation(
            confidence_pct, float(cv_mae or 312.66), total_net_sales, feature_contributions
        )

        is_test = str(form_data.get('is_test', '')).lower() in {'1', 'true', 'on', 'yes'}
        event_name = (form_data.get('event_name') or '').strip()
        if not event_name:
            industry_label = form_data.get('industry', 'Event')
            city_label = form_data.get('city', 'Unknown')
            event_name = f"{industry_label} - {city_label} ({event_date_str})"
        model_id = f"franchise_{franchise_id}" if model_type == 'franchise_specific' else 'clean_layers_ensemble'
        prediction_id = uuid.uuid4().hex
        actual_total = None
        actual_str = (form_data.get('actual_total_net_sales') or '').strip()
        if actual_str:
            try:
                actual_total = float(actual_str)
            except ValueError:
                actual_total = None
        submitted_status = 'booked_confirmed' if str(form_data.get('is_booked', '')).lower() in {'1', 'true', 'on', 'yes'} else 'predicted_only'
        if actual_total is not None:
            submitted_status = 'completed'

        saved = franchise_db.record_prediction(
            prediction_id=prediction_id,
            franchise_id=franchise_id,
            model_id=model_id,
            event_name=event_name,
            duration_hours=duration,
            predicted_revenue_per_hour=float(net_sales_per_hour),
            predicted_total_revenue=float(total_net_sales),
            confidence_lower=float(total_lower),
            confidence_upper=float(total_upper),
            is_test=is_test,
            actual_total_net_sales=actual_total,
            event_status=submitted_status,
            include_in_training=actual_total is not None,
            scheduled_event_date=event_date_str,
            event_features_json=_serialize_feature_snapshot(feature_df),
        )
        if not saved:
            print("[WARN] Failed to save prediction record")
        
        # Return results in format expected by frontend
        franchise = franchise_db.get_franchise(franchise_id)
        target_per_hour = None
        if franchise:
            target_per_hour = franchise.get('target_net_sales_per_hour')

        return jsonify({
            'success': True,
            'net_sales_per_hour': float(net_sales_per_hour),
            'total_net_sales': float(total_net_sales),
            'per_hour_lower': float(per_hour_lower),
            'per_hour_upper': float(per_hour_upper),
            'range_lower': float(total_lower),
            'range_upper': float(total_upper),
            'confidence_percentage': float(confidence_pct),
            'model_cv_mae': float(cv_mae or 312.66),
            'model_type': model_type,
            'event_status': submitted_status,
            'duration': f"{duration:.1f}",
            'target_net_sales_per_hour': float(target_per_hour) if target_per_hour else None,
            # Event details for summary display
            'event_details': {
                'industry': form_data.get('industry', 'Unknown'),
                'city': form_data.get('city', 'Unknown'),
                'zip_code': form_data.get('zip_code', 'Unknown'),
                'date': event_date_str,
                'time': form_data.get('start_time', 'Unknown'),
                'duration': f"{duration:.1f}",
            },
            # Demographics for display
            'demographics': {
                'median_income': f"${demographics.get('Median_Household_Income', 0):,.0f}",
                'population': f"{demographics.get('Total_Population', 0):,.0f}",
                'unemployment_rate': f"{demographics.get('Unemployment_Rate', 0):.1f}%",
                'source': 'census_api' if demographics.get('source') == 'census_api' else 'default',
            },
            'feature_contributions': feature_contributions,
            'confidence_explanations': confidence_explanations,
            'advanced_explanation': advanced_explanation,
            'explanation': confidence_explanations,
            'features_used': list(feature_df.columns),
        })
        
    except Exception as e:
        print(f"Prediction error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'An internal error has occurred. Please try again later.'
        }), 500


# ===== REST OF ROUTES (Dashboard, Login, etc) =====


@app.route('/api/industries')
@require_auth
def get_industries():
    """Return list of available industries from the training data."""
    # Get industries from the historical data
    industries = INDUSTRIES
    return jsonify({'industries': industries})


# ===== EQUIPMENT CONFIG ROUTES =====

@app.route('/equipment-config')
@require_auth
def equipment_config_page():
    """Render the equipment configuration page."""
    return render_template('equipment_config.html')


@app.route('/api/equipment-mappings', methods=['GET'])
@require_auth
def get_equipment_mappings():
    """Return all equipment mappings for the current franchise."""
    franchise_id = get_current_franchise_id()
    mappings = franchise_db.get_equipment_mappings(franchise_id)
    return jsonify({'mappings': mappings})


@app.route('/api/equipment-mappings', methods=['POST'])
@require_auth
def upsert_equipment_mapping():
    """Add or update an equipment mapping."""
    franchise_id = get_current_franchise_id()
    data = request.get_json(silent=True) or {}
    equipment_name = str(data.get('equipment_name') or '').strip()
    equipment_type = str(data.get('equipment_type') or '').strip()
    notes = str(data.get('notes') or '').strip()
    if not equipment_name or not equipment_type:
        return jsonify({'error': 'equipment_name and equipment_type are required'}), 400
    ok = franchise_db.upsert_equipment_mapping(franchise_id, equipment_name, equipment_type, notes)
    if not ok:
        return jsonify({'error': 'Failed to save equipment mapping'}), 500
    return jsonify({'status': 'ok'})


@app.route('/api/equipment-mappings/<path:equipment_name>', methods=['DELETE'])
@require_auth
def delete_equipment_mapping(equipment_name: str):
    """Delete an equipment mapping by name."""
    franchise_id = get_current_franchise_id()
    franchise_db.delete_equipment_mapping(franchise_id, equipment_name)
    return jsonify({'status': 'ok'})


# ===== BULK UPLOAD ROUTES =====

# Date/time formats used by the scheduling software exports.
_SCHEDULING_DATETIME_FORMATS = [
    '%b %d, %Y %I:%M %p',   # Apr 04, 2026 08:15 AM
    '%m/%d/%Y %I:%M %p',    # 04/04/2026 08:15 AM
    '%Y-%m-%d %H:%M:%S',
    '%Y-%m-%d %H:%M',
]


def _parse_scheduling_dt(value: object):
    """Parse a datetime from a scheduling export cell (string or datetime)."""
    if isinstance(value, datetime):
        return value
    text = str(value or '').strip()
    if not text or text == '-':
        return None
    for fmt in _SCHEDULING_DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _scheduling_duration_hours(start, end) -> str:
    """Return duration in hours between two datetime objects, or empty string."""
    if start is None or end is None:
        return ''
    delta = end - start
    hours = delta.total_seconds() / 3600
    return str(round(hours, 2)) if hours > 0 else ''


def _scheduling_zip(value: object) -> str:
    """Normalize ZIP code — strip trailing .0 from float-ish values."""
    text = str(value or '').strip()
    if re.match(r'^\d+\.0$', text):
        return text[:text.index('.')]
    return text


def _convert_scheduling_export(file_bytes) -> 'pd.DataFrame | None':
    """Detect a scheduling software export and return a bulk-upload-ready DataFrame.

    Supports two formats:
      - ``-events``: clean header at row 1 with 'Event Code' as the first column.
      - ``-report``: 'Sales History Report' summary format; data table begins at
        the row where column A == 'Event Name'.

    Returns ``None`` if the file does not look like either scheduling format.
    """
    try:
        import openpyxl as _opxl
    except ImportError:
        return None

    try:
        file_bytes.seek(0)
        wb = _opxl.load_workbook(file_bytes, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None

    if not rows:
        return None

    first_cell = str(rows[0][0] or '').strip()

    payment_term_headers = {'payment term', 'payment terms'}

    # --- -events format ---
    if first_cell == 'Event Code':
        headers = [str(h or '').strip() for h in rows[0]]

        def _payment_term_is_menu(row_dict: dict) -> bool:
            for key, value in row_dict.items():
                if str(key or '').strip().lower() in payment_term_headers:
                    return str(value or '').strip().lower() == 'menu'
            return True

        records = []
        for row in rows[1:]:
            if not any(v is not None for v in row):
                continue
            d = dict(zip(headers, row))
            if not _payment_term_is_menu(d):
                continue
            start = _parse_scheduling_dt(d.get('Event Start Date Time'))
            end = _parse_scheduling_dt(d.get('Event End Date Time'))
            records.append({
                'Event Name': str(d.get('Event Name') or '').strip(),
                'Industry': str(d.get('Event Industry') or '').strip(),
                'Equipment Type': '',
                'Equipment Number': str(d.get('Event Equipment') or '').strip(),
                'Source Event ID': str(d.get('Event Code') or '').strip(),
                'Event Date': start.strftime('%Y-%m-%d') if start else '',
                'Start Time': start.strftime('%H:%M') if start else '',
                'Duration (Hours)': _scheduling_duration_hours(start, end),
                'City': str(d.get('City') or '').strip(),
                'ZIP Code': _scheduling_zip(d.get('Zip Code')),
                'Temperature (°F)': '',
                'Net Sales': d.get('Event Sales Collected ($)') or '',
            })
        return pd.DataFrame(records) if records else None

    # --- -report format ---
    if first_cell == 'Sales History Report':
        header_idx = None
        for i, row in enumerate(rows):
            if str(row[0] or '').strip() == 'Event Name':
                header_idx = i
                break
        if header_idx is None:
            return None
        headers = [str(h or '').strip() for h in rows[header_idx]]

        def _payment_term_is_menu(row_dict: dict) -> bool:
            for key, value in row_dict.items():
                if str(key or '').strip().lower() in payment_term_headers:
                    return str(value or '').strip().lower() == 'menu'
            return True

        records = []
        for row in rows[header_idx + 1:]:
            if not any(v is not None for v in row):
                continue
            d = dict(zip(headers, row))
            if not _payment_term_is_menu(d):
                continue
            start = _parse_scheduling_dt(d.get('Event Start Date'))
            end = _parse_scheduling_dt(d.get('Event End Date'))
            records.append({
                'Event Name': str(d.get('Event Name') or '').strip(),
                'Industry': str(d.get('Industry') or '').strip(),
                'Equipment Type': '',
                'Equipment Number': '',
                'Source Event ID': '',
                'Event Date': start.strftime('%Y-%m-%d') if start else '',
                'Start Time': start.strftime('%H:%M') if start else '',
                'Duration (Hours)': _scheduling_duration_hours(start, end),
                'City': str(d.get('City') or '').strip(),
                'ZIP Code': _scheduling_zip(d.get('ZipCode')),
                'Temperature (°F)': '',
                'Net Sales': d.get('Sales $') or '',
            })
        return pd.DataFrame(records) if records else None

    return None


BULK_UPLOAD_COLUMNS = [
    'Event Name',
    'Industry',
    'Equipment Type',
    'Equipment Number',
    'Source Event ID',
    'Event Date',
    'Start Time',
    'Duration (Hours)',
    'City',
    'ZIP Code',
    'Temperature (°F)',
    'Net Sales',
]

BULK_UPLOAD_EXAMPLE_ROW = [
    'Summer Festival',
    'Festival',
    'Blended Truck',
    'TRK-12',
    'sched-evt-2025-08-15-001',
    '2025-08-15',
    '10:00',
    '6',
    'Portland',
    '97201',
    '75',
    '',
]


def _normalize_bulk_key_text(value: object) -> str:
    """Normalize text so dedup keys are case/whitespace/punctuation insensitive."""
    text = str(value or '').strip().lower()
    if not text:
        return ''
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^a-z0-9 ]+', '', text)
    return text.strip()


def _normalize_bulk_time(value: object) -> str:
    """Normalize start time to HH:MM when possible."""
    text = str(value or '').strip()
    if not text:
        return ''
    for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p', '%I:%M%p'):
        try:
            return datetime.strptime(text, fmt).strftime('%H:%M')
        except ValueError:
            continue
    return _normalize_bulk_key_text(text)


def _build_bulk_strong_dedup_key(
    source_event_id: str,
    equipment_number: str,
    event_date_str: str,
    start_time_str: str,
    duration_hours: float,
    city: str,
    zip_code: str,
) -> str:
    """
    Build a durable dedup key only when we have a stable identifier.
    Without stable IDs, we avoid cross-upload dedup to prevent false positives.
    """
    source_id_norm = _normalize_bulk_key_text(source_event_id)
    if source_id_norm:
        return f'v2:source:{source_id_norm}'

    equipment_norm = _normalize_bulk_key_text(equipment_number)
    if equipment_norm:
        return '|'.join([
            'v2:eq',
            _normalize_bulk_key_text(event_date_str),
            _normalize_bulk_time(start_time_str),
            f'{float(duration_hours):.3f}',
            _normalize_bulk_key_text(city),
            _normalize_bulk_key_text(zip_code),
            equipment_norm,
        ])

    return ''


def _build_bulk_fallback_dedup_key(
    event_name: str,
    industry: str,
    event_date_str: str,
    start_time_str: str,
    duration_hours: float,
    city: str,
    zip_code: str,
) -> str:
    """Build a conservative fallback dedup key for schedule imports without stable IDs."""
    return '|'.join([
        'v2:weak',
        _normalize_bulk_key_text(event_name),
        _normalize_bulk_key_text(industry),
        _normalize_bulk_key_text(event_date_str),
        _normalize_bulk_time(start_time_str),
        f'{float(duration_hours):.3f}',
        _normalize_bulk_key_text(city),
        _normalize_bulk_key_text(zip_code),
    ])


def _build_bulk_row_fingerprint(parts: list[str]) -> str:
    """Hash full row content to catch accidental duplicate rows in the same file."""
    normalized = '|'.join(_normalize_bulk_key_text(part) for part in parts)
    return hashlib.sha1(normalized.encode('utf-8')).hexdigest()


@app.route('/bulk-upload/template')
@require_auth
def bulk_upload_template():
    """Return a downloadable Excel template for bulk event upload."""
    import io
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl is not installed on the server'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Events'
    ws.append(BULK_UPLOAD_COLUMNS)
    ws.append(BULK_UPLOAD_EXAMPLE_ROW)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = (
        'attachment; filename="bulk_upload_template.xlsx"'
    )
    return response


@app.route('/bulk-upload', methods=['GET', 'POST'])
@require_auth
def bulk_upload_page():
    """Render and process the bulk event upload page."""
    franchise_id = get_current_franchise_id()
    recent_uploads = franchise_db.get_recent_batch_uploads(franchise_id, limit=10)

    if request.method == 'GET':
        return render_template('bulk_upload.html', recent_uploads=recent_uploads)

    import io as _io

    def _filter_menu_payment_terms(df_in: 'pd.DataFrame'):
        """If Payment Term columns exist, keep only menu rows and return skipped count."""
        payment_term_col = next(
            (
                col
                for col in df_in.columns
                if str(col).strip().lower() in {'payment term', 'payment terms'}
            ),
            None,
        )
        if payment_term_col is None:
            return df_in, 0

        payment_series = df_in[payment_term_col].fillna('').astype(str).str.strip().str.lower()
        menu_mask = payment_series.eq('menu')
        ignored_count = int((~menu_mask).sum())
        return df_in[menu_mask].copy(), ignored_count

    ignored_non_menu_rows = 0

    # ------------------------------------------------------------------
    # Phase 2: user classified unknown equipment and is resubmitting.
    # ------------------------------------------------------------------
    pending_token = request.form.get('pending_file_token', '').strip()
    if pending_token:
        if not re.match(r'^[0-9a-f]{32}$', pending_token):
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Invalid pending upload token.']},
            )

        pending_dir = os.path.realpath(os.path.join(tempfile.gettempdir(), 'kona_pending_uploads'))
        franchise_id_safe = str(franchise_id)
        safe_name_re = re.compile(rf'^([0-9a-f]{{32}})_{re.escape(franchise_id_safe)}\.tmp$')
        expected_name = None
        try:
            for entry in os.listdir(pending_dir):
                m = safe_name_re.match(entry)
                if m and m.group(1) == pending_token:
                    expected_name = entry
                    break
        except OSError:
            expected_name = None

        if not expected_name:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Invalid pending upload token.']},
            )

        candidate_path = os.path.join(pending_dir, expected_name)
        pending_path = os.path.realpath(candidate_path)
        expected_path = os.path.join(pending_dir, expected_name)
        if os.path.commonpath([pending_dir, pending_path]) != pending_dir or pending_path != expected_path:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Invalid pending upload token.']},
            )
        if os.path.basename(pending_path) != expected_name:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Invalid pending upload token.']},
            )
        if not os.path.exists(pending_path):
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Pending upload expired or not found. Please re-upload the file.']},
            )

        # Save the user-supplied equipment type mappings before processing.
        for key, value in request.form.items():
            if key.startswith('equip_') and value:
                equip_name = key[len('equip_'):]
                franchise_db.upsert_equipment_mapping(franchise_id, equip_name, value, '')

        try:
            if not os.path.isfile(pending_path):
                raise OSError("Pending upload is not a regular file.")
            open_flags = os.O_RDONLY
            if hasattr(os, 'O_NOFOLLOW'):
                open_flags |= os.O_NOFOLLOW
            fd = os.open(pending_path, open_flags)
            with os.fdopen(fd, 'rb') as fh:
                file_bytes = _io.BytesIO(fh.read())
            uploaded_filename_for_record = request.form.get('pending_original_filename', 'upload.xlsx')
        except OSError:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Could not read pending upload. Please re-upload.']},
            )
        finally:
            try:
                os.remove(pending_path)
            except OSError:
                pass

        converted_df = _convert_scheduling_export(file_bytes)
        if converted_df is not None:
            df = converted_df
        else:
            file_bytes.seek(0)
            try:
                df = pd.read_excel(file_bytes, engine='openpyxl', dtype=str)
            except Exception as exc:
                return render_template(
                    'bulk_upload.html',
                    recent_uploads=recent_uploads,
                    results={'added': 0, 'outcomes': 0, 'skipped': 0,
                             'errors': [f'Could not read file: {exc}']},
                )

        df, ignored_non_menu_rows = _filter_menu_payment_terms(df)

    # ------------------------------------------------------------------
    # Phase 1: fresh file upload.
    # ------------------------------------------------------------------
    else:
        uploaded_file = request.files.get('file')
        if not uploaded_file or not uploaded_file.filename:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['No file was uploaded.']},
            )

        filename = uploaded_file.filename.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': ['Only .xlsx and .xls files are accepted.']},
            )

        try:
            file_bytes = _io.BytesIO(uploaded_file.read())
            converted_df = _convert_scheduling_export(file_bytes)
            if converted_df is not None:
                df = converted_df
            else:
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes, engine='openpyxl', dtype=str)
        except Exception as exc:
            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                results={'added': 0, 'outcomes': 0, 'skipped': 0,
                         'errors': [f'Could not read file: {exc}']},
            )

        df, ignored_non_menu_rows = _filter_menu_payment_terms(df)

        uploaded_filename_for_record = uploaded_file.filename

        # Detect equipment numbers in the file that have no franchise mapping.
        if 'Equipment Number' in df.columns:
            known_mappings = {
                m['equipment_name'].strip().lower()
                for m in franchise_db.get_equipment_mappings(franchise_id)
            }
            unknown_equipment = sorted({
                str(v).strip()
                for v in df['Equipment Number'].dropna()
                if str(v).strip() and str(v).strip().lower() not in known_mappings
            })
        else:
            unknown_equipment = []

        if unknown_equipment:
            pending_dir = os.path.join(tempfile.gettempdir(), 'kona_pending_uploads')
            os.makedirs(pending_dir, exist_ok=True)
            pending_token = uuid.uuid4().hex
            pending_path = os.path.join(pending_dir, f'{pending_token}_{franchise_id}.tmp')
            file_bytes.seek(0)
            with open(pending_path, 'wb') as fh:
                fh.write(file_bytes.read())

            return render_template(
                'bulk_upload.html',
                recent_uploads=recent_uploads,
                unknown_equipment=unknown_equipment,
                pending_file_token=pending_token,
                pending_original_filename=uploaded_file.filename,
            )

    required_cols = {
        'Industry',
        'Event Date',
        'Start Time',
        'Duration (Hours)',
        'City',
        'ZIP Code',
    }
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        return render_template(
            'bulk_upload.html',
            recent_uploads=recent_uploads,
            results={'added': 0, 'outcomes': 0, 'skipped': 0,
                     'errors': [f'Missing required columns: {", ".join(sorted(missing_cols))}']},
        )

    # Ensure the bulk_dedup_key column exists (added once, idempotent via PRAGMA check)
    import sqlite3 as _sqlite3
    try:
        conn = franchise_db.get_connection()
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(predictions)")
        col_names = {row_info[1] for row_info in cur.fetchall()}
        if 'bulk_dedup_key' not in col_names:
            cur.execute("ALTER TABLE predictions ADD COLUMN bulk_dedup_key TEXT DEFAULT ''")
            conn.commit()
        conn.close()
    except _sqlite3.OperationalError:
        pass

    # Build lookup for existing dedup keys for this franchise.
    existing_predictions_by_key = {}
    try:
        conn = franchise_db.get_connection()
        cur = conn.cursor()
        cur.execute(
            '''SELECT prediction_id, bulk_dedup_key, actual_total_net_sales, event_features_json
               FROM predictions
               WHERE franchise_id = ?
                 AND bulk_dedup_key IS NOT NULL
                 AND TRIM(bulk_dedup_key) != ''
            ''',
            (franchise_id,)
        )
        for row in cur.fetchall():
            key = row['bulk_dedup_key']
            if isinstance(key, str) and key:
                existing_predictions_by_key[key] = {
                    'prediction_id': row['prediction_id'],
                    'actual_total_net_sales': row['actual_total_net_sales'],
                    'has_feature_snapshot': bool(row['event_features_json'] and str(row['event_features_json']).strip()),
                }
        conn.close()
    except Exception:
        existing_predictions_by_key = {}

    seen_file_fingerprints = set()

    added = 0
    outcomes = 0
    skipped = ignored_non_menu_rows
    errors = []
    today = datetime.utcnow().date()
    batch_model_id = 'bulk_upload'

    def _cell_text(row_data, column_name: str) -> str:
        value = row_data.get(column_name, '')
        if pd.isna(value):
            return ''
        return str(value).strip()

    if PROD_MANAGER is None:
        return render_template(
            'bulk_upload.html',
            recent_uploads=recent_uploads,
            results={'added': 0, 'outcomes': 0, 'skipped': 0,
                     'errors': ['Model not loaded. Please try again later.']},
        )

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # 1-based, header is row 1

        event_name = _cell_text(row, 'Event Name')
        industry = _cell_text(row, 'Industry')
        equipment_type = _cell_text(row, 'Equipment Type')
        equipment_number = _cell_text(row, 'Equipment Number')
        source_event_id = _cell_text(row, 'Source Event ID')
        event_date_str = _cell_text(row, 'Event Date')
        start_time_str = _cell_text(row, 'Start Time')
        duration_str = _cell_text(row, 'Duration (Hours)')
        city = _cell_text(row, 'City')
        zip_code = _cell_text(row, 'ZIP Code')
        temperature_str = _cell_text(row, 'Temperature (°F)')
        net_sales_str = _cell_text(row, 'Net Sales')

        # Infer equipment_type from equipment_number via franchise DB mapping when blank.
        if not equipment_type and equipment_number:
            inferred = franchise_db.lookup_equipment_type(franchise_id, equipment_number)
            if inferred:
                equipment_type = inferred

        # Validate required fields
        row_errors = []
        if not event_date_str:
            row_errors.append('Event Date is required')
        if not city:
            row_errors.append('City is required')
        if not zip_code:
            row_errors.append('ZIP Code is required')
        if not duration_str:
            row_errors.append('Duration (Hours) is required')
        if not start_time_str:
            row_errors.append('Start Time is required')
        if not industry:
            row_errors.append('Industry is required')

        if row_errors:
            errors.append(f'Row {row_num}: {"; ".join(row_errors)}')
            continue

        # Parse event date
        try:
            event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
        except ValueError:
            errors.append(f'Row {row_num}: Invalid Event Date format (expected YYYY-MM-DD)')
            continue

        # Parse duration
        try:
            duration_hours = float(duration_str)
            if duration_hours <= 0:
                raise ValueError('non-positive')
        except ValueError:
            errors.append(f'Row {row_num}: Invalid Duration (Hours) value')
            continue

        row_fingerprint = _build_bulk_row_fingerprint([
            event_name,
            industry,
            equipment_type,
            equipment_number,
            source_event_id,
            event_date_str,
            start_time_str,
            duration_str,
            city,
            zip_code,
            net_sales_str,
        ])

        # Always dedup exact duplicate rows within the same uploaded file.
        if row_fingerprint in seen_file_fingerprints:
            skipped += 1
            continue

        strong_dedup_key = _build_bulk_strong_dedup_key(
            source_event_id=source_event_id,
            equipment_number=equipment_number,
            event_date_str=event_date_str,
            start_time_str=start_time_str,
            duration_hours=duration_hours,
            city=city,
            zip_code=zip_code,
        )
        dedup_key = strong_dedup_key or _build_bulk_fallback_dedup_key(
            event_name=event_name,
            industry=industry,
            event_date_str=event_date_str,
            start_time_str=start_time_str,
            duration_hours=duration_hours,
            city=city,
            zip_code=zip_code,
        )

        # Parse optional net sales before dedup handling so existing-event
        # updates can apply outcomes from this upload.
        net_sales = None
        if net_sales_str:
            try:
                net_sales = float(net_sales_str)
                if net_sales < 0:
                    net_sales = None
            except ValueError:
                pass

        existing = existing_predictions_by_key.get(dedup_key)
        existing_prediction_id_for_backfill = None
        if existing:
            updated_existing_outcome = False
            if event_date < today and net_sales and net_sales > 0 and existing.get('actual_total_net_sales') is None:
                updated_existing_outcome = franchise_db.update_actual_net_sales(
                    franchise_id,
                    existing['prediction_id'],
                    net_sales,
                )
            if updated_existing_outcome:
                outcomes += 1
            if existing.get('has_feature_snapshot'):
                skipped += 1
                seen_file_fingerprints.add(row_fingerprint)
                continue

            # Existing record is missing feature snapshot. Recompute features and
            # backfill the existing row instead of inserting a duplicate.
            existing_prediction_id_for_backfill = existing.get('prediction_id')

        # Build event_name if blank
        if not event_name:
            event_name = f'{industry or "Event"} - {city} ({event_date_str})'

        prediction_id = uuid.uuid4().hex

        # Generate prediction for each uploaded event.
        try:
            zip_code_norm = str(zip_code).zfill(5)
            prediction_form_data = {
                'industry': industry,
                'equipment_type': equipment_type,
                'event_date': event_date_str,
                'start_time': start_time_str,
                'duration': str(duration_hours),
                'city': city,
                'zip_code': zip_code_norm,
                'temperature': temperature_str,
                'has_smoothie_capability': '1' if equipment_type.lower() == 'blended truck' else '0',
            }

            demographics = get_demographics_from_zip(zip_code_norm)
            weather_info = get_weather_for_event(
                zip_code_norm,
                datetime.strptime(event_date_str, '%Y-%m-%d'),
                user_temperature=temperature_str,
            )

            from clean_layers_feature_preparation import prepare_features_for_clean_layers_model
            feature_df, _ = prepare_features_for_clean_layers_model(
                form_data=prediction_form_data,
                franchise_id=franchise_id,
                demographics_dict=demographics,
                weather_dict=weather_info,
                franchise_history=None,
                return_raw_values=True,
            )

            if feature_df is None or feature_df.empty:
                errors.append(f'Row {row_num}: Could not prepare model features for prediction')
                continue

            predictions, model_type = PROD_MANAGER.predict(feature_df, franchise_id=franchise_id)
            y_hat = float(predictions[0]) if hasattr(predictions, '__len__') else float(predictions)
            total_net_sales = y_hat
            net_sales_per_hour = total_net_sales / duration_hours if duration_hours > 0 else total_net_sales
            total_lower = max(0.0, total_net_sales * 0.8)
            total_upper = total_net_sales * 1.2
            model_id = f'franchise_{franchise_id}' if model_type == 'franchise_specific' else 'clean_layers_ensemble'
            batch_model_id = model_id
        except Exception as pred_err:
            errors.append(f'Row {row_num}: Prediction failed: {pred_err}')
            continue

        if existing_prediction_id_for_backfill:
            try:
                conn = franchise_db.get_connection()
                cur = conn.cursor()
                cur.execute(
                    '''
                    UPDATE predictions
                    SET event_features_json = ?,
                        include_in_training = CASE
                            WHEN actual_total_net_sales IS NOT NULL THEN 1
                            ELSE include_in_training
                        END
                    WHERE prediction_id = ? AND franchise_id = ?
                    ''',
                    (
                        _serialize_feature_snapshot(feature_df),
                        existing_prediction_id_for_backfill,
                        franchise_id,
                    ),
                )
                conn.commit()
                conn.close()
            except Exception as backfill_err:
                errors.append(f'Row {row_num}: Could not backfill feature snapshot: {backfill_err}')

            skipped += 1
            seen_file_fingerprints.add(row_fingerprint)
            existing_predictions_by_key[dedup_key]['has_feature_snapshot'] = True
            continue

        if event_date < today:
            row_status = 'completed' if net_sales and net_sales > 0 else 'needs_outcome'
        else:
            row_status = 'booked_confirmed'

        saved = franchise_db.record_prediction(
            prediction_id=prediction_id,
            franchise_id=franchise_id,
            model_id=model_id,
            event_name=event_name,
            duration_hours=duration_hours,
            predicted_revenue_per_hour=float(net_sales_per_hour),
            predicted_total_revenue=float(total_net_sales),
            confidence_lower=float(total_lower),
            confidence_upper=float(total_upper),
            is_test=False,
            actual_total_net_sales=None,
            event_status=row_status,
            include_in_training=False,
            scheduled_event_date=event_date_str,
            event_features_json=_serialize_feature_snapshot(feature_df),
        )
        if not saved:
            errors.append(f'Row {row_num}: Could not save prediction record')
            continue

        # Persist dedup key so future uploads can detect duplicates
        try:
            conn = franchise_db.get_connection()
            cur = conn.cursor()
            cur.execute(
                "UPDATE predictions SET bulk_dedup_key = ? WHERE prediction_id = ?",
                (dedup_key, prediction_id),
            )
            conn.commit()
            conn.close()
        except _sqlite3.OperationalError as _db_err:
            errors.append(f'Row {row_num}: Could not store dedup key: {_db_err}')

        seen_file_fingerprints.add(row_fingerprint)
        existing_predictions_by_key[dedup_key] = {
            'prediction_id': prediction_id,
            'actual_total_net_sales': None,
            'has_feature_snapshot': True,
        }
        added += 1

        # If past event with positive net sales, log the outcome
        if event_date < today and net_sales and net_sales > 0:
            updated = franchise_db.update_actual_net_sales(franchise_id, prediction_id, net_sales)
            if updated:
                outcomes += 1
                existing_predictions_by_key[dedup_key]['actual_total_net_sales'] = net_sales

    # Record the batch upload in history
    batch_id = uuid.uuid4().hex
    franchise_db.record_batch_upload(
        batch_id=batch_id,
        franchise_id=franchise_id,
        model_id=batch_model_id,
        original_filename=uploaded_filename_for_record,
        event_count=added,
        result_csv_path='',
    )

    recent_uploads = franchise_db.get_recent_batch_uploads(franchise_id, limit=10)
    return render_template(
        'bulk_upload.html',
        recent_uploads=recent_uploads,
        results={'added': added, 'outcomes': outcomes, 'skipped': skipped, 'errors': errors},
    )


if __name__ == '__main__':
    # Create templates directory if it doesn't exist
    os.makedirs('templates', exist_ok=True)
    
    print("="*80)
    print("Event Revenue Prediction Web App")
    print("="*80)
    model_name = type(model).__name__
    print(f"Model: {model_name}")
    if cv_mae is not None and cv_std is not None:
        try:
            print(f"CV MAE: {float(cv_mae):.2f} ± {float(cv_std):.2f}")
        except Exception:
            pass
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    print(f"Starting server on http://127.0.0.1:5000 (debug={debug_mode})")
    print("="*80)
    
    app.run(debug=debug_mode, host='127.0.0.1', port=5000)
