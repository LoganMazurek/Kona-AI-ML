"""Convert scheduling software exports to the Kona ML bulk upload format.

Supported source formats
------------------------
  -events  (``Events_<date>-events.xlsx``)
      Clean header row at row 1.  Contains Event Code, equipment, full sales
      detail.  **Preferred format** — use this when available.

  -report  (``EventSalesReport_<date>-report.xlsx``)
      Formatted "Sales History Report" with summary rows at the top.  Data
      table begins at the first row whose first cell is "Event Name".  Lacks
      Source Event ID and equipment columns.

Usage
-----
    python scripts/scheduling_export_converter.py path/to/input.xlsx
    python scripts/scheduling_export_converter.py path/to/input.xlsx -o path/to/output.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BULK_UPLOAD_COLUMNS = [
    "Event Name",
    "Industry",
    "Equipment Type",
    "Equipment Number",
    "Source Event ID",
    "Event Date",
    "Start Time",
    "Duration (Hours)",
    "City",
    "ZIP Code",
    "Temperature (°F)",
    "Net Sales",
]

# Date/time formats used by the scheduling software.
_DATETIME_FORMATS = [
    "%b %d, %Y %I:%M %p",   # Apr 04, 2026 08:15 AM
    "%m/%d/%Y %I:%M %p",    # 04/04/2026 08:15 AM
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
]

_PAYMENT_TERM_COLUMNS = {
    "payment term",
    "payment terms",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_dt(value: object) -> datetime | None:
    """Parse a datetime value from either a Python datetime or a string."""
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text or text == "-":
        return None
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _duration_hours(start: datetime | None, end: datetime | None) -> str:
    """Return duration in hours rounded to 2 decimal places, or empty string."""
    if start is None or end is None:
        return ""
    delta = end - start
    hours = delta.total_seconds() / 3600
    if hours <= 0:
        return ""
    return str(round(hours, 2))


def _zip(value: object) -> str:
    """Format ZIP code — remove trailing .0 from float-ish values."""
    text = str(value or "").strip()
    if re.match(r"^\d+\.0$", text):
        return text[: text.index(".")]
    return text


def _has_payment_term_column(headers: list[object]) -> bool:
    """Return True when the source sheet includes a payment-term column."""
    return any(str(h or "").strip().lower() in _PAYMENT_TERM_COLUMNS for h in headers)


def _payment_term_value(row: dict[object, object]) -> str:
    """Read payment term from a row using case-insensitive known column names."""
    for key, value in row.items():
        if str(key or "").strip().lower() in _PAYMENT_TERM_COLUMNS:
            return str(value or "").strip().lower()
    return ""


# ---------------------------------------------------------------------------
# Format detection & parsing
# ---------------------------------------------------------------------------

def _detect_format(ws) -> str:
    """Return 'events' or 'report' based on the first non-empty cell."""
    for row in ws.iter_rows(max_row=5, values_only=True):
        first = str(row[0] or "").strip() if row else ""
        if first == "Event Code":
            return "events"
        if first in ("Sales History Report", "Event Name", "From Date"):
            return "report"
    return "unknown"


def _read_events_format(ws) -> pd.DataFrame:
    """Parse the -events format (header at row 1)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Worksheet is empty.")

    headers = [str(h or "").strip() for h in rows[0]]
    has_payment_term = _has_payment_term_column(headers)
    data = [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]

    records = []
    for d in data:
        if has_payment_term and _payment_term_value(d) != "menu":
            continue

        start = _parse_dt(d.get("Event Start Date Time"))
        end = _parse_dt(d.get("Event End Date Time"))

        records.append(
            {
                "Event Name": str(d.get("Event Name") or "").strip(),
                "Industry": str(d.get("Event Industry") or "").strip(),
                "Equipment Type": "",
                "Equipment Number": str(d.get("Event Equipment") or "").strip(),
                "Source Event ID": str(d.get("Event Code") or "").strip(),
                "Event Date": start.strftime("%Y-%m-%d") if start else "",
                "Start Time": start.strftime("%H:%M") if start else "",
                "Duration (Hours)": _duration_hours(start, end),
                "City": str(d.get("City") or "").strip(),
                "ZIP Code": _zip(d.get("Zip Code")),
                "Temperature (°F)": "",
                "Net Sales": d.get("Event Sales Collected ($)") or "",
            }
        )

    return pd.DataFrame(records, columns=BULK_UPLOAD_COLUMNS)


def _read_report_format(ws) -> pd.DataFrame:
    """Parse the -report format (data table starts after summary rows)."""
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row — first row where col[0] == "Event Name"
    header_idx = None
    for i, row in enumerate(rows):
        if str(row[0] or "").strip() == "Event Name":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            "Could not find 'Event Name' header row in -report file. "
            "Is this a Sales History Report export?"
        )

    headers = [str(h or "").strip() for h in rows[header_idx]]
    has_payment_term = _has_payment_term_column(headers)
    data = [
        dict(zip(headers, row))
        for row in rows[header_idx + 1 :]
        if any(v is not None for v in row)
    ]

    records = []
    for d in data:
        if has_payment_term and _payment_term_value(d) != "menu":
            continue

        start = _parse_dt(d.get("Event Start Date"))
        end = _parse_dt(d.get("Event End Date"))

        records.append(
            {
                "Event Name": str(d.get("Event Name") or "").strip(),
                "Industry": str(d.get("Industry") or "").strip(),
                "Equipment Type": "",
                "Equipment Number": "",
                "Source Event ID": "",
                "Event Date": start.strftime("%Y-%m-%d") if start else "",
                "Start Time": start.strftime("%H:%M") if start else "",
                "Duration (Hours)": _duration_hours(start, end),
                "City": str(d.get("City") or "").strip(),
                "ZIP Code": _zip(d.get("ZipCode")),
                "Temperature (°F)": "",
                "Net Sales": d.get("Sales $") or "",
            }
        )

    return pd.DataFrame(records, columns=BULK_UPLOAD_COLUMNS)


# ---------------------------------------------------------------------------
# Main conversion entry point
# ---------------------------------------------------------------------------

def convert(input_path: str | Path, output_path: str | Path | None = None) -> Path:
    """Convert a scheduling export to a bulk upload xlsx.

    Parameters
    ----------
    input_path:
        Path to the source .xlsx file.
    output_path:
        Destination path.  Defaults to ``<input_stem>_bulk_upload.xlsx``
        in the same directory as the input file.

    Returns
    -------
    Path
        The path of the generated output file.
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_bulk_upload.xlsx"
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    fmt = _detect_format(ws)

    if fmt == "events":
        df = _read_events_format(ws)
        print(f"Detected format: events  ({len(df)} rows)")
    elif fmt == "report":
        df = _read_report_format(ws)
        print(f"Detected format: report  ({len(df)} rows)")
    else:
        raise ValueError(
            f"Unrecognized file format in '{input_path.name}'. "
            "Expected a -events or -report export from the scheduling software."
        )

    wb.close()

    # Write output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Bulk Upload"
    out_ws.append(BULK_UPLOAD_COLUMNS)
    for _, row in df.iterrows():
        out_ws.append([row[col] for col in BULK_UPLOAD_COLUMNS])

    out_wb.save(output_path)
    print(f"Saved: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert a scheduling software export to the Kona ML bulk upload format."
    )
    parser.add_argument("input", help="Path to the source .xlsx file (-events or -report).")
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output path. Defaults to <input_stem>_bulk_upload.xlsx.",
    )
    args = parser.parse_args()

    try:
        convert(args.input, args.output)
    except (ValueError, FileNotFoundError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
